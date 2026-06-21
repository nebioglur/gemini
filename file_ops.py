# -*- coding: utf-8 -*-
import pandas as pd
from tkinter import filedialog, messagebox
from datetime import datetime
import os
from openpyxl.styles import Font, PatternFill, Alignment
import re

def export_tree_to_excel(tree, analiz_detaylari, initial_file="Veriler.xlsx", export_columns=None, use_reg_date=False, custom_summary=None, filter_tag=None):
    try:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=initial_file, filetypes=[("Excel Files", "*.xlsx")])
        if not file_path: return

        # --- ÖZET HESAPLAMA ---
        if custom_summary:
            summary_str = custom_summary
        else:
            cnt_uyumsuz = 0
            cnt_dikkat = 0
            cnt_uyumlu = 0
            cnt_f_uyumsuz = 0
            for val in analiz_detaylari.values():
                durum = str(val.get('durum', '')).upper()
                if "F/UYUMSUZ" in durum: cnt_f_uyumsuz += 1
                elif val.get('alarm') == "KIRMIZI": cnt_uyumsuz += 1
                elif val.get('alarm') == "SARI": cnt_dikkat += 1
                elif val.get('alarm') == "YESIL": cnt_uyumlu += 1
            
            summary_str = f"ÖZET: UYUMSUZ: {cnt_uyumsuz} | F/UYUMSUZ: {cnt_f_uyumsuz} | DİKKAT: {cnt_dikkat} | UYUMLU: {cnt_uyumlu}"

        data = []
        date_col_header = "KAYIT TAR." if use_reg_date else "RASAT TAR."
        base_cols = ["TÜRÜ", "KULL.", date_col_header, "BÜLTEN"]
        if filter_tag == "LATE_ARRIVAL":
            base_cols.append("GECİKME SÜRESİ")
        
        def traverse(parent=""):
            for child in tree.get_children(parent):
                if filter_tag:
                    tags = tree.item(child, "tags")
                    if isinstance(filter_tag, (list, tuple)):
                        if not any(t in tags for t in filter_tag):
                            traverse(child)
                            continue
                    else:
                        if filter_tag not in tags:
                            traverse(child)
                            continue

                text = tree.item(child)["text"]
                vals = tree.item(child)["values"]

                if vals: 
                    # vals indeksleri: 0:KULL, 1:GÖND, 2:KAYIT, 3:RASAT, 4:BÜLTEN, 5:UYUM
                    date_val = vals[2] if use_reg_date else vals[3]
                    row = [text, vals[0], date_val, vals[4]]
                    
                    if filter_tag == "LATE_ARRIVAL":
                        delay_str = ""
                        if len(vals) > 5:
                            match = re.search(r'\((\d+)\s*dk\)', str(vals[5]))
                            if match: delay_str = f"{match.group(1)} dk"
                        row.append(delay_str)

                    data.append(row)
                traverse(child)
        traverse()
        
        df = pd.DataFrame(data, columns=base_cols)
        
        if export_columns:
            cols_to_keep = [c for c in export_columns if c in df.columns]
            df = df[cols_to_keep]
            
        # A4 ve Yazdırma Ayarları ile Kaydet
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Veriyi 2. satırdan başlat (1. satıra özet gelecek)
            df.to_excel(writer, index=False, sheet_name='Rapor', startrow=1)
            ws = writer.sheets['Rapor']
            
            # Özet Bilgisi (A1 Hücresi)
            ws['A1'] = summary_str
            ws['A1'].font = Font(bold=True, size=11, color="FF0000")

            ws.page_setup.paperSize = 9  # A4
            ws.page_setup.orientation = 'landscape'
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToHeight = False
            ws.page_setup.fitToWidth = 1
            
            # Tüm hücreler için Metni Kaydır (Wrap Text) ve Üste Hizala
            align_style = Alignment(wrap_text=True, vertical='top', horizontal='left')
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = align_style

            # Başlık Satırı Stili (Gri Arka Plan + Kalın)
            # Başlık artık 2. satırda
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            for cell in ws[2]:
                cell.font = header_font
                cell.fill = header_fill

            # TAF Satırlarını Kalın Yap
            bold_font = Font(bold=True)
            turu_idx = -1
            bulten_idx = -1
            # DataFrame sütunlarına göre indeksleri bul
            for i, col in enumerate(df.columns):
                if col == "TÜRÜ": turu_idx = i
                elif col == "BÜLTEN": bulten_idx = i
            
            # Veri satırlarını gez (Başlık hariç, 3. satırdan başla)
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                is_taf = False
                # Hücre değerlerini kontrol et (row[idx] hücresidir)
                if turu_idx != -1 and row[turu_idx].value and "TAF" in str(row[turu_idx].value):
                    is_taf = True
                elif bulten_idx != -1 and row[bulten_idx].value and str(row[bulten_idx].value).strip().startswith("TAF"):
                    is_taf = True
                
                if is_taf:
                    for cell in row: cell.font = bold_font

            # Sütun Genişliklerini A4 Yatay İçin Optimize Et
            # Başlık satırını (2. satır) kontrol ederek genişlikleri ata
            for cell in ws[2]:
                col_letter = cell.column_letter
                val = str(cell.value).upper()
                if "BÜLTEN" in val:
                    ws.column_dimensions[col_letter].width = 85  # Bülten için geniş alan (Wrap ile aşağı uzar)
                elif "RASAT TAR" in val or "KAYIT TAR" in val:
                    ws.column_dimensions[col_letter].width = 18
                elif "TÜRÜ" in val:
                    ws.column_dimensions[col_letter].width = 35  # TAF başlıkları sığsın
                elif "KULL" in val:
                    ws.column_dimensions[col_letter].width = 12
                elif "GECİKME" in val:
                    ws.column_dimensions[col_letter].width = 15
                else:
                    ws.column_dimensions[col_letter].width = 15

        messagebox.showinfo("Başarılı", "Veriler Excel'e aktarıldı.")
    except Exception as e:
        messagebox.showerror("Hata", f"Excel'e aktarım hatası: {e}")

def save_detailed_log_file(tree, analiz_detaylari, initial_file="Analiz_Log.txt", filter_tag=None):
    if not analiz_detaylari:
        messagebox.showwarning("Uyarı", "Kaydedilecek analiz verisi yok.")
        return
        
    try:
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            initialfile=initial_file,
            filetypes=[("Metin Dosyası", "*.txt"), ("Log Dosyası", "*.log")],
            title="Detaylı Log Kaydet"
        )
        if not file_path: return
        
        # --- ÖZET HESAPLAMA ---
        cnt_uyumsuz = 0
        cnt_dikkat = 0
        cnt_uyumlu = 0
        cnt_f_uyumsuz = 0
        for val in analiz_detaylari.values():
            durum = str(val.get('durum', '')).upper()
            if "F/UYUMSUZ" in durum: cnt_f_uyumsuz += 1
            elif val.get('alarm') == "KIRMIZI": cnt_uyumsuz += 1
            elif val.get('alarm') == "SARI": cnt_dikkat += 1
            elif val.get('alarm') == "YESIL": cnt_uyumlu += 1

        with open(file_path, "w", encoding="utf-8") as f:
            f.write(f"KARDELEN DETAYLI ANALİZ RAPORU\n")
            f.write(f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n")
            f.write(f"ÖZET: UYUMSUZ: {cnt_uyumsuz} | F/UYUMSUZ: {cnt_f_uyumsuz} | DİKKAT: {cnt_dikkat} | UYUMLU: {cnt_uyumlu}\n")
            f.write("="*80 + "\n\n")
            
            def write_detail(iid):
                if filter_tag:
                    tags = tree.item(iid, "tags")
                    if isinstance(filter_tag, (list, tuple)):
                        if not any(t in tags for t in filter_tag):
                            return
                    else:
                        if filter_tag not in tags:
                            return

                if iid in analiz_detaylari:
                    detay = analiz_detaylari[iid]
                    
                    # Kullanıcı bilgisini (KULL.) çekmeye çalış
                    kull_bilgisi = ""
                    try:
                        vals = tree.item(iid)['values']
                        if vals: kull_bilgisi = f"(Kull: {vals[0]}) "
                    except: pass

                    durum = "BİLİNMİYOR"
                    if detay['alarm'] == "YESIL": durum = "UYUMLU"
                    elif detay['alarm'] == "SARI": durum = "DIKKAT"
                    elif detay['alarm'] == "KIRMIZI": durum = "UYUMSUZ"
                    durum = detay.get('durum', 'BİLİNMİYOR').upper()
                    if "RE/GECMIS HATASI" in durum: durum = "RE/GEÇMİŞ HATASI"
                    elif detay['alarm'] == "YESIL": durum = "UYUMLU"
                    elif detay['alarm'] == "SARI": durum = "DİKKAT"
                    elif detay['alarm'] == "KIRMIZI" and "F/UYUMSUZ" not in durum: durum = "UYUMSUZ"
                    elif detay['alarm'] == "MAVI": durum = "DÜZELTİLDİ"

                    f.write(f"[{detay['tarih']}] {kull_bilgisi}{detay['metar']}\n")
                    f.write(f"   DURUM: {durum}\n")
                    ref_taf = detay.get('taf', '')
                    clean_ref_taf = ref_taf.split('=')[0] + '=' if '=' in ref_taf else ref_taf
                    f.write(f"   ► REF TAF: {clean_ref_taf}\n")
                    
                    # Düzeltilmiş (AAA vb.) TAF işlemi varsa farkını/nedenini de rapora yazdır
                    if durum == "DÜZELTİLDİ" and detay.get("reasons"):
                        for r in detay["reasons"]:
                            f.write(f"   ℹ️ {r}\n")
                            
                    f.write("\n")

            for item_id in tree.get_children():
                if item_id not in analiz_detaylari:
                    vals = tree.item(item_id)['values']
                    text = tree.item(item_id)['text']
                    if "TAF" in text:
                            f.write(f"\n>>> TAF GRUBU: {text} (Yayın: {vals[2]})\n")
                            f.write("-" * 50 + "\n")
                    
                    # Alt öğeleri (rasatları) işle
                    for child_id in tree.get_children(item_id):
                        write_detail(child_id)
                else:
                    # TAF'sız (yetim) rasat
                    write_detail(item_id)
        
        messagebox.showinfo("Başarılı", f"Log dosyası kaydedildi:\n{file_path}")
    except Exception as e:
        messagebox.showerror("Hata", f"Log kaydetme hatası: {e}")

def export_map_history_excel(df, parent_window, filter_val):
    try:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], parent=parent_window)
        if not file_path: return

        df_filtered = df.copy()
        if filter_val == "❌ UYUMSUZ":
            df_filtered = df_filtered[df_filtered["_uyum"].astype(str).str.contains("UYUMSUZ", na=False)]
        elif filter_val == "⚠️ DİKKAT":
            df_filtered = df_filtered[df_filtered["_uyum"].astype(str).str.contains("DİKKAT", na=False)]
        elif filter_val == "🔸 TREND YOK":
            df_filtered = df_filtered[df_filtered["_uyum"].astype(str).str.contains("TREND YOK", na=False)]
        
        export_cols = ["date", "Türü", "İstasyon", "Bülten", "_uyum", "_detay", "_ref_taf"]
        existing_cols = [c for c in export_cols if c in df_filtered.columns]
        df_export = df_filtered[existing_cols]
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='Geçmiş')
            ws = writer.sheets['Geçmiş']
            
            ws.page_setup.paperSize = 9  # A4
            ws.page_setup.orientation = 'landscape'
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToHeight = False
            ws.page_setup.fitToWidth = 1
            
            # Tüm hücreler için Metni Kaydır (Wrap Text) ve Üste Hizala
            align_style = Alignment(wrap_text=True, vertical='top', horizontal='left')
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = align_style

            # Başlık Satırı Stili (Gri Arka Plan + Kalın)
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill

            # TAF Satırlarını Kalın Yap
            bold_font = Font(bold=True)
            turu_idx = -1
            bulten_idx = -1
            
            for i, col in enumerate(df_export.columns):
                if col == "Türü": turu_idx = i
                elif col == "Bülten": bulten_idx = i
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                is_taf = False
                if turu_idx != -1 and row[turu_idx].value and "TAF" in str(row[turu_idx].value).upper():
                    is_taf = True
                elif bulten_idx != -1 and row[bulten_idx].value and str(row[bulten_idx].value).strip().upper().startswith("TAF"):
                    is_taf = True
                
                if is_taf:
                    for cell in row: cell.font = bold_font

            for column in ws.columns:
                column_letter = column[0].column_letter
                col_name = str(column[0].value)
                if "Bülten" in col_name:
                    ws.column_dimensions[column_letter].width = 85
                elif "date" in col_name or "TAR" in col_name:
                    ws.column_dimensions[column_letter].width = 18
                elif "Türü" in col_name:
                    ws.column_dimensions[column_letter].width = 15
                elif "_detay" in col_name:
                    ws.column_dimensions[column_letter].width = 50
                else:
                    ws.column_dimensions[column_letter].width = 15

        messagebox.showinfo("Başarılı", "Geçmiş veriler Excel'e aktarıldı.", parent=parent_window)
    except Exception as e:
        messagebox.showerror("Hata", f"Excel aktarım hatası: {e}", parent=parent_window)

def export_map_scores_excel(scores_list, parent_window):
    try:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], parent=parent_window)
        if not file_path: return
        
        df = pd.DataFrame(scores_list)
        df = df.rename(columns={"code": "İstasyon", "score": "Puan", "count": "Rasat Sayısı"})
        df = df.sort_values(by=["Puan", "Rasat Sayısı"], ascending=False)
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Skorlar')
            ws = writer.sheets['Skorlar']
            
            ws.page_setup.paperSize = 9  # A4
            ws.page_setup.orientation = 'portrait' # Skor tablosu dar olduğu için Dikey
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToHeight = False
            ws.page_setup.fitToWidth = 1
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                adjusted_width = (max_length + 2)
                if adjusted_width > 50: adjusted_width = 50
                ws.column_dimensions[column_letter].width = adjusted_width

        messagebox.showinfo("Başarılı", "Skor tablosu Excel'e aktarıldı.", parent=parent_window)
    except Exception as e:
        messagebox.showerror("Hata", f"Excel aktarım hatası: {e}", parent=parent_window)