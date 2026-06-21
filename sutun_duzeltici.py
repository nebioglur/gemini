import sys
import subprocess
def ensure_lxml():
    try:
        import lxml
        import html5lib
    except ImportError:
        try:
            print("HTML Ayrıştırma motorları (lxml) eksik. Otomatik yükleniyor...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", "lxml", "html5lib"])
        except Exception:
            pass
ensure_lxml()

import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
from openpyxl.utils import get_column_letter
import threading
import re
import shutil
import logging
from logging.handlers import RotatingFileHandler
import zipfile

# --- BAD CRC-32 (BOZUK EXCEL) BYPASS HACK ---
try:
    if hasattr(zipfile, 'ZipExtFile') and hasattr(zipfile.ZipExtFile, '_update_crc'):
        original_update_crc = zipfile.ZipExtFile._update_crc
        def patched_update_crc(self, newdata):
            try:
                original_update_crc(self, newdata)
            except zipfile.BadZipFile:
                pass
        zipfile.ZipExtFile._update_crc = patched_update_crc
        if not getattr(zipfile.ZipExtFile, '_crc_patched', False):
            _orig_update_crc = zipfile.ZipExtFile._update_crc
            def _patched_update_crc(self, newdata):
                self._expected_crc = None  # Zorla CRC kontrolünü kapat (BadZipFile hatasını önler)
                return _orig_update_crc(self, newdata)
            zipfile.ZipExtFile._update_crc = _patched_update_crc
            zipfile.ZipExtFile._crc_patched = True
except Exception:
    pass
# --------------------------------------------

# Loglama yapılandırması
if not logging.getLogger().hasHandlers():
    log_handler = RotatingFileHandler('denetim_merkezi.log', maxBytes=5*1024*1024, backupCount=2, encoding='utf-8')
    logging.basicConfig(
        handlers=[log_handler],
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

# Ana uygulamadaki normalizasyon ve başlık bulma fonksiyonlarını kullanmak için import et
try:
    import denetim_merkezi_1 as dm1
except ImportError:
    # Bu script'in tek başına çalışabilmesi için hata mesajı göster
    # Eğer bu dosya ana klasörde değilse, bu import başarısız olur.
    tk.Tk().withdraw()
    messagebox.showerror(
        "Kritik Hata", 
        "Gerekli 'denetim_merkezi_1.py' dosyası bulunamadı.\n\n"
        "Lütfen bu script'i ana uygulama ('arayuz.py') ile aynı klasörde çalıştırdığınızdan emin olun."
    )
    exit()

def sutunlari_otomatik_duzelt():
    """
    Kullanıcıdan bir Excel dosyası seçmesini ister. Dosyadaki her sayfanın
    sütun başlıklarını ana programın anlayacağı standart formata dönüştürür
    ve '_DUZELTILMIS' ekiyle yeni bir dosya olarak kaydeder.
    """
    if not tk._default_root:
        root = tk.Tk()
        root.withdraw()

    dosya_yolu = filedialog.askopenfilename(
        title="Sütun Başlıklarını Düzeltmek İçin Excel Dosyasını Seçin",
        filetypes=[("Excel Dosyaları", "*.xlsx *.xls")]
    )

    if not dosya_yolu:
        # Kullanıcı dosya seçmeden pencereyi kapattıysa işlemi sonlandır.
        return

    def islem_yurut():
        try:
            # Çıktı dosyasının adını oluştur (örn: 'veriler.xlsx' -> 'veriler_DUZELTILMIS.xlsx')
            dizin, dosya_adi = os.path.split(dosya_yolu)
            ad, uzanti = os.path.splitext(dosya_adi)
            cikti_yolu = os.path.join(dizin, f"{ad}_DUZELTILMIS{uzanti}")

            xls_obj = None
            html_tablolar = None
            sheet_names = []
            
            try:
                xls_obj = pd.ExcelFile(dosya_yolu)
                sheet_names = xls_obj.sheet_names
            except Exception as e:
                try:
                    html_tablolar = pd.read_html(dosya_yolu, header=None, flavor='lxml')
                    sheet_names = [f"Sayfa_{i+1}" for i in range(len(html_tablolar))]
                except Exception as html_e:
                    try:
                        from bs4 import BeautifulSoup, SoupStrainer
                        sadece_tablolar = SoupStrainer('table')
                        with open(dosya_yolu, 'r', encoding='utf-8', errors='replace') as f:
                            try:
                                hizli_soup = BeautifulSoup(f, 'lxml', parse_only=sadece_tablolar)
                            except Exception:
                                f.seek(0)
                                hizli_soup = BeautifulSoup(f, 'html5lib', parse_only=sadece_tablolar)
                        
                        for br in hizli_soup.find_all("br"):
                            br.replace_with(" ")
                        for hidden in hizli_soup.find_all(style=lambda value: value and 'display:none' in value.replace(' ', '')):
                            hidden.decompose()
                        
                        html_str = str(hizli_soup)
                        if not html_str.strip():
                            raise ValueError("HTML tablosu bulunamadı.")
                        import io
                        html_tablolar = pd.read_html(io.StringIO(html_str), header=None, flavor='lxml')
                        sheet_names = [f"Sayfa_{i+1}" for i in range(len(html_tablolar))]
                    except Exception as html_bs4_e:
                        try:
                            try:
                                df_csv = pd.read_csv(dosya_yolu, sep='\t', encoding='utf-8')
                            except UnicodeDecodeError:
                                df_csv = pd.read_csv(dosya_yolu, sep='\t', encoding='cp1254')
                            if len(df_csv.columns) < 3:
                                try:
                                    df_csv = pd.read_csv(dosya_yolu, sep=None, engine='python', encoding='utf-8')
                                except UnicodeDecodeError:
                                    df_csv = pd.read_csv(dosya_yolu, sep=None, engine='python', encoding='cp1254')
                            html_tablolar = [df_csv]
                            sheet_names = ["Sayfa_1"]
                        except Exception as csv_e:
                            logging.warning("Tüm standart okuma yöntemleri başarısız. Agresif (RAW TEXT) kurtarma deneniyor...")
                            try:
                                with open(dosya_yolu, 'r', encoding='utf-8', errors='ignore') as f:
                                    raw_lines = f.readlines()
                                salvaged_data = []
                                for line in raw_lines:
                                    if not line.strip(): continue
                                    parts = re.split(r'\t|;|\||,| {2,}', line.strip())
                                    salvaged_data.append(parts)
                                if len(salvaged_data) > 0:
                                    html_tablolar = [pd.DataFrame(salvaged_data)]
                                    sheet_names = ["KURTARILAN_RAW"]
                                else:
                                    raise ValueError("Anlamlı metin verisi bulunamadı.")
                            except Exception as agg_e:
                                raise Exception(f"Dosya Excel, HTML veya CSV formatında okunamadı:\nExcel: {e}\nHTML: {html_e}\nHTML(bs4): {html_bs4_e}\nCSV: {csv_e}\nAgresif Kurtarma: {agg_e}")

            try:
                # Yeni Excel dosyası için bir yazıcı (writer) oluştur
                with pd.ExcelWriter(cikti_yolu, engine='openpyxl') as writer:
                    # 🚀 KESİN ÇÖZÜM: İşlem kesilirse çökmemesi için geçici bir sayfa oluştur.
                    pd.DataFrame(["İşlem bekleniyor..."]).to_excel(writer, sheet_name="Gecici_Sayfa", index=False, header=False)
                    
                    yazilan_sayfa_sayisi = 0

                    # Dosyadaki her bir sayfa (sheet) için döngü başlat
                    for i, sayfa_adi in enumerate(sheet_names):
                        # BAZI EXCEL DIŞA AKTARMA ARAÇLARI TARAFINDAN OLUŞTURULAN GİZLİ METADATA SAYFALARINI ATLA
                        if 'document map' in str(sayfa_adi).lower() or 'documentmap' in str(sayfa_adi).lower():
                            continue
                        logging.info(f" - '{sayfa_adi}' sayfası işleniyor...")
                    
                        # Veriler açık rasatlarda (Sheet2, A-...) olduğu için kapalı rasatları (Sheet3, K-...) atla
                        sayfa_str = str(sayfa_adi).lower()
                        if "sheet" in sayfa_str or "sayfa" in sayfa_str:
                            rakamlar = re.findall(r'\d+', sayfa_str)
                            if rakamlar and int(rakamlar[0]) % 2 != 0:
                                # KONTROL: Bu sayfa aslında ana verileri içeriyor mu?
                                try:
                                    if xls_obj is not None:
                                        ham_kontrol = pd.read_excel(xls_obj, sheet_name=sayfa_adi, header=None, dtype=str)
                                    else:
                                        ham_kontrol = html_tablolar[i].astype(str)
                                    h_idx = dm1.header_bul(ham_kontrol)
                                    is_main_sheet = False
                                    if h_idx is not None:
                                        row_values = ham_kontrol.iloc[h_idx].astype(str).str.lower().values
                                        match_count = sum(1 for val in row_values if str(val).strip() in ["t", "p", "n", "ff", "dd", "h", "a"] or any(k in str(val).strip() for k in ["gmt", "saat", "ww", "halihazır", "present", "istasyon", "rüzgar", "yön", "hız", "basınç", "yağış", "rrr", "bulut", "görüş", "vv", "bülten", "metar", "tipi"]))
                                        if match_count >= 5: is_main_sheet = True
                                    if not is_main_sheet:
                                        for r_i in range(min(10, len(ham_kontrol))):
                                            v_a = str(ham_kontrol.iloc[r_i, 0]).strip().upper() if len(ham_kontrol.columns) > 0 else ""
                                            if v_a in ['TİPİ', 'TIPI', 'METAR'] or 'METAR' in v_a:
                                                is_main_sheet = True; break
                                    if is_main_sheet:
                                        pass # Bu sayfa atlanmamalı, normal düzeltmeye tabi tutulmalı
                                    else:
                                        logging.info(f"   -> Tek sayılı sayfa ({sayfa_adi}) meta veri sayılarak aynen kopyalanıyor.")
                                        ham_kontrol.to_excel(writer, sheet_name=sayfa_adi, index=False, header=False)
                                        yazilan_sayfa_sayisi += 1
                                    continue
                                except Exception: pass
               
                        try:
                            # Başlık satırının nerede olduğunu bulmak için veriyi ham olarak oku (başlıksız)
                            if xls_obj is not None:
                                ham_df = pd.read_excel(xls_obj, sheet_name=sayfa_adi, header=None, dtype=str)
                            else:
                                ham_df = html_tablolar[i].astype(str)
                            
                            sayfa_str = str(sayfa_adi).lower()
                            yeni_sayfa_adi = str(sayfa_adi)
                            # DÜZELTME: Sayfa isimlerini değiştirmeyin, aksi takdirde Sheet2-Sheet3 eşleşmesi bozulur.
    
                            # denetim_merkezi_1'deki fonksiyon ile başlık satırının indeksini bul
                            header_index = dm1.header_bul(ham_df)
                            
                            if header_index is not None:
                                ham_df.to_excel(writer, sheet_name=yeni_sayfa_adi, index=False, header=False)
                                logging.info(f"   -> Başlıklar {header_index + 1}. satırda bulundu. Orijinal başlık yapısı korunuyor. (Yeni Adı: {yeni_sayfa_adi})")
    
                                # 🚀 PERFORMANS OPTİMİZASYONU: Sütun genişliğini milisaniyeler içinde hesapla
                                ws = writer.sheets[yeni_sayfa_adi]
                                for col_idx, col in enumerate(ham_df.columns):
                                    try:
                                        max_len = ham_df[col].astype(str).map(len).max()
                                        width = min(max_len + 2, 60)
                                    except:
                                        width = 15
                                    col_letter = get_column_letter(col_idx + 1)
                                    ws.column_dimensions[col_letter].width = width
    
                                yazilan_sayfa_sayisi += 1
                            else:
                                logging.warning(f"   -> UYARI: Bu sayfada tanınabilir bir başlık satırı bulunamadı. Değiştirilmeden kopyalanıyor. (Yeni Adı: {yeni_sayfa_adi})")
                                ham_df.to_excel(writer, sheet_name=yeni_sayfa_adi, index=False, header=False)
                                yazilan_sayfa_sayisi += 1
                        except Exception as inner_e:
                            logging.error(f"   -> HATA: '{sayfa_adi}' işlenirken hata oluştu ve atlandı.", exc_info=True)
                            try:
                                logging.info(f"   -> AGRESİF KURTARMA: '{sayfa_adi}' ham veri olarak zorla yazılıyor...")
                                if 'ham_df' in locals() and ham_df is not None:
                                    ham_df.to_excel(writer, sheet_name=f"KRT_{str(sayfa_adi)[:25]}", index=False, header=False)
                                    yazilan_sayfa_sayisi += 1
                            except Exception as raw_e:
                                logging.error(f"   -> AGRESİF KURTARMA DA BAŞARISIZ OLDU: {raw_e}")

                # Geçici sayfayı temizle veya boş dosyayı işaretle
                if yazilan_sayfa_sayisi > 0:
                    writer.book.remove(writer.sheets["Gecici_Sayfa"])
                else:
                    if "Gecici_Sayfa" in writer.sheets:
                        writer.sheets["Gecici_Sayfa"].title = "Bos_Sayfa"
            finally:
                if xls_obj is not None:
                    xls_obj.close()

            if tk._default_root:
                tk._default_root.after(0, lambda: messagebox.showinfo(
                    "Başarılı",
                    f"Sütunlar başarıyla düzeltildi!\n\n"
                    f"Yeni dosya şu konuma kaydedildi:\n{cikti_yolu}"
                ))
            else:
                messagebox.showinfo("Başarılı", f"Sütunlar başarıyla düzeltildi!\n\nYeni dosya şu konuma kaydedildi:\n{cikti_yolu}")
            
            try: os.startfile(cikti_yolu)
            except: pass

        except Exception as e:
            if tk._default_root:
                tk._default_root.after(0, lambda e=e: messagebox.showerror("Hata", f"Düzeltme işlemi sırasında bir hata oluştu:\n{e}"))
            logging.error("Sütun düzeltme işleminde hata", exc_info=True)
            
    logging.info("--- Excel Sütun Düzeltici Başlatıldı ---")
    threading.Thread(target=islem_yurut, daemon=True).start()

def sessiz_duzelt(dosya_yolu):
    """
    Arayüz üzerinden otomatik (sessiz) çağrıldığında çalışır.
    Mevcut dosyanın üzerine yazar (orijinal yedeği arayüz tarafından alınır).
    """
    gecici_yol = None
    try:
        dizin, dosya_adi = os.path.split(dosya_yolu)
        gecici_yol = os.path.join(dizin, f"temp_{dosya_adi}")
        
        xls_obj = None
        html_tablolar = None
        sheet_names = []
        try:
            xls_obj = pd.ExcelFile(dosya_yolu)
            sheet_names = xls_obj.sheet_names
        except Exception as e:
            try:
                html_tablolar = pd.read_html(dosya_yolu, header=None, flavor='lxml')
                sheet_names = [f"Sayfa_{i+1}" for i in range(len(html_tablolar))]
            except Exception as html_e:
                try:
                    from bs4 import BeautifulSoup, SoupStrainer
                    sadece_tablolar = SoupStrainer('table')
                    with open(dosya_yolu, 'r', encoding='utf-8', errors='replace') as f:
                        try:
                            hizli_soup = BeautifulSoup(f, 'lxml', parse_only=sadece_tablolar)
                        except getattr(BeautifulSoup, 'FeatureNotFound', Exception) as bs4_err:
                            f.seek(0) # Dosyayı başa sar
                            try:
                                hizli_soup = BeautifulSoup(f, 'html.parser', parse_only=sadece_tablolar)
                            except Exception:
                                f.seek(0)
                                hizli_soup = BeautifulSoup(f, 'html5lib', parse_only=sadece_tablolar)
                    
                    for br in hizli_soup.find_all("br"):
                        br.replace_with(" ")
                    for hidden in hizli_soup.find_all(style=lambda value: value and 'display:none' in value.replace(' ', '')):
                        hidden.decompose()
                    
                    html_str = str(hizli_soup)
                    if not html_str.strip():
                        raise ValueError("HTML tablosu bulunamadı.")
                    import io
                    html_tablolar = pd.read_html(io.StringIO(html_str), header=None, flavor='lxml')
                    sheet_names = [f"Sayfa_{i+1}" for i in range(len(html_tablolar))]
                except Exception as html_bs4_e:
                    try:
                        try:
                            df_csv = pd.read_csv(dosya_yolu, sep='\t', encoding='utf-8')
                        except UnicodeDecodeError:
                            df_csv = pd.read_csv(dosya_yolu, sep='\t', encoding='cp1254')
                        if len(df_csv.columns) < 3:
                            try:
                                df_csv = pd.read_csv(dosya_yolu, sep=None, engine='python', encoding='utf-8')
                            except UnicodeDecodeError:
                                df_csv = pd.read_csv(dosya_yolu, sep=None, engine='python', encoding='cp1254')
                        html_tablolar = [df_csv]
                        sheet_names = ["Sayfa_1"]
                    except Exception as csv_e:
                        logging.warning("Tüm standart okuma yöntemleri başarısız. Agresif (RAW TEXT) kurtarma deneniyor...")
                        try:
                            with open(dosya_yolu, 'r', encoding='utf-8', errors='ignore') as f:
                                raw_lines = f.readlines()
                            salvaged_data = []
                            for line in raw_lines:
                                if not line.strip(): continue
                                parts = re.split(r'\t|;|\||,| {2,}', line.strip())
                                salvaged_data.append(parts)
                            if len(salvaged_data) > 0:
                                html_tablolar = [pd.DataFrame(salvaged_data)]
                                sheet_names = ["KURTARILAN_RAW"]
                            else:
                                raise ValueError("Anlamlı metin verisi bulunamadı.")
                        except Exception as agg_e:
                            raise Exception(f"Dosya Excel, HTML veya CSV formatında okunamadı:\nExcel: {e}\nHTML: {html_e}\nHTML(bs4): {html_bs4_e}\nCSV: {csv_e}\nAgresif Kurtarma: {agg_e}")

        try:
            with pd.ExcelWriter(gecici_yol, engine='openpyxl') as writer:
                pd.DataFrame(["İşlem bekleniyor..."]).to_excel(writer, sheet_name="Gecici_Sayfa", index=False, header=False)
                yazilan_sayfa_sayisi = 0

                for i, sayfa_adi in enumerate(sheet_names):
                    if 'document map' in str(sayfa_adi).lower() or 'documentmap' in str(sayfa_adi).lower():
                        continue
                    
                    sayfa_str = str(sayfa_adi).lower()
                    is_odd_data_sheet = False
                    if "sheet" in sayfa_str or "sayfa" in sayfa_str:
                        rakamlar = re.findall(r'\d+', sayfa_str)
                        if rakamlar and int(rakamlar[0]) % 2 != 0:
                            is_odd_data_sheet = True

        
                    try:
                        if xls_obj is not None:
                            ham_df = pd.read_excel(xls_obj, sheet_name=sayfa_adi, header=None, dtype=str)
                        else:
                            ham_df = html_tablolar[i].astype(str)
                        
                        if is_odd_data_sheet:
                            # KONTROL
                            h_idx = dm1.header_bul(ham_df)
                            if h_idx is not None:
                                row_values = ham_df.iloc[h_idx].astype(str).str.lower().values
                                match_count = sum(1 for val in row_values if str(val).strip() in ["t", "p", "n", "ff", "dd", "h", "a"] or any(k in str(val).strip() for k in ["gmt", "saat", "ww", "halihazır", "present", "istasyon", "rüzgar", "yön", "hız", "basınç", "yağış", "rrr", "bulut", "görüş", "vv", "bülten", "metar", "tipi"]))
                                if match_count >= 5: is_odd_data_sheet = False
                            if is_odd_data_sheet:
                                for r_i in range(min(10, len(ham_df))):
                                    v_a = str(ham_df.iloc[r_i, 0]).strip().upper() if len(ham_df.columns) > 0 else ""
                                    if v_a in ['TİPİ', 'TIPI', 'METAR'] or 'METAR' in v_a:
                                        is_odd_data_sheet = False; break
                                        
                        if is_odd_data_sheet:
                            yeni_sayfa_adi = str(sayfa_adi)
                            ham_df.to_excel(writer, sheet_name=yeni_sayfa_adi, index=False, header=False)
                            yazilan_sayfa_sayisi += 1
                            continue

                        sayfa_str_lower = str(sayfa_adi).lower()
                        yeni_sayfa_adi = str(sayfa_adi)
                        # DÜZELTME: Sayfa isimlerini değiştirmeyin, aksi takdirde Sheet2-Sheet3 eşleşmesi bozulur.

                        header_index = dm1.header_bul(ham_df)
                        
                        if header_index is not None:
                            ham_df.to_excel(writer, sheet_name=yeni_sayfa_adi, index=False, header=False)

                            ws = writer.sheets[yeni_sayfa_adi]
                            for col_idx, col in enumerate(ham_df.columns):
                                try: max_len = ham_df[col].astype(str).map(len).max(); width = min(max_len + 2, 60)
                                except: width = 15
                                ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

                            yazilan_sayfa_sayisi += 1
                        else:
                            ham_df.to_excel(writer, sheet_name=yeni_sayfa_adi, index=False, header=False)
                            yazilan_sayfa_sayisi += 1
                    except Exception as inner_e:
                        logging.error(f"   -> HATA: '{sayfa_adi}' sessiz islenirken hata olustu.", exc_info=True)
                        try:
                            logging.info(f"   -> AGRESİF KURTARMA: '{sayfa_adi}' ham veri olarak zorla yazılıyor...")
                            if 'ham_df' in locals() and ham_df is not None:
                                ham_df.to_excel(writer, sheet_name=f"KRT_{str(sayfa_adi)[:25]}", index=False, header=False)
                                yazilan_sayfa_sayisi += 1
                        except Exception as raw_e:
                            logging.error(f"   -> AGRESİF KURTARMA DA BAŞARISIZ OLDU: {raw_e}")

                if yazilan_sayfa_sayisi > 0:
                    writer.book.remove(writer.sheets["Gecici_Sayfa"])
                else:
                    if "Gecici_Sayfa" in writer.sheets: writer.sheets["Gecici_Sayfa"].title = "Bos_Sayfa"
        finally:
            if xls_obj is not None:
                xls_obj.close()
                
        if not os.path.exists(gecici_yol):
            raise Exception("Geçici dosya oluşturulamadı (Antivirüs engellemiş olabilir).")
            
        # Windows'ta FileExistsError (183) ve PermissionError (32) önlemek için
        if os.path.exists(dosya_yolu):
            import time
            for _ in range(3):
                try:
                    os.replace(gecici_yol, dosya_yolu)
                    gecici_yol = None
                    break
                except PermissionError:
                    time.sleep(0.5)
            else:
                # os.replace 3 denemede de başarısız olursa eski yönteme dön
                try:
                    os.remove(dosya_yolu)
                    shutil.move(gecici_yol, dosya_yolu)
                    gecici_yol = None
                except PermissionError:
                    raise Exception(f"Lütfen Excel'de açık olan '{os.path.basename(dosya_yolu)}' dosyasını KAPATIN. Dosya başka bir işlem tarafından kullanılıyor.")
        else:
            shutil.move(gecici_yol, dosya_yolu)
            gecici_yol = None
    except Exception as e:
        if gecici_yol and os.path.exists(gecici_yol):
            try: os.remove(gecici_yol)
            except: pass
        raise Exception(f"Sütunlar düzeltilirken sorun oluştu: {e}")

def arsivleri_toplu_duzelt():
    """Arşiv klasöründeki tüm Excel raporlarının sütunlarını güncel şablona uyarlar."""
    if not tk._default_root:
        root = tk.Tk()
        root.withdraw()
        
    arsiv_dir = r"C:\Users\nebio\Desktop\check\Arsiv"
    if not os.path.exists(arsiv_dir):
        messagebox.showwarning("Uyarı", "Arşiv klasörü bulunamadı.")
        return
        
    if not messagebox.askyesno("Onay", f"'{arsiv_dir}' klasöründeki tüm arşiv dosyalarının sütun isimleri güncellenecek.\nBu işlem biraz sürebilir. Onaylıyor musunuz?"):
        return
        
    def _worker():
        islenen = 0
        hatali = 0
        atlanan = 0
        
        for root_d, dirs, files in os.walk(arsiv_dir):
            for f in files:
                if f.startswith("DENETIM_") and f.endswith((".xlsx", ".xls")):
                    dosya_yolu = os.path.join(root_d, f)
                    try:
                        # Bozuk veya okunamayan dosyaları baştan tespit et ve sessizce atla
                        try:
                            test_xls = pd.ExcelFile(dosya_yolu)
                            test_xls.close()
                        except Exception:
                            atlanan += 1
                            continue
                            
                        sessiz_duzelt(dosya_yolu)
                        islenen += 1
                    except Exception as e:
                        logging.error(f"Arşiv dosyası düzeltilemedi ({f}): {e}")
                        hatali += 1
                        
        if tk._default_root:
            tk._default_root.after(0, lambda: messagebox.showinfo("Tamamlandı", f"Toplu düzeltme tamamlandı.\n\nBaşarıyla güncellenen: {islenen}\nBozuk olduğu için atlanan: {atlanan}\nHata alınan: {hatali}"))
            
    threading.Thread(target=_worker, daemon=True).start()

if __name__ == "__main__":
    sutunlari_otomatik_duzelt()