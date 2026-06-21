import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import os
import kurallar
import threading
import re
import math
import datetime
import calendar

def dosya_oku_html_veya_excel(dosya_yolu, header_satiri):
    tum_kayitlar = []

    try:
        # 1️⃣ Önce normal Excel dene
        excel = pd.ExcelFile(dosya_yolu)
        sayfalar = excel.sheet_names

        for sayfa in sayfalar:
            df = pd.read_excel(excel, sheet_name=sayfa, header=header_satiri)
            if not df.empty:
                df["sayfa"] = sayfa
                tum_kayitlar.append(df)

    except Exception as e:
        # 2️⃣ Excel değilse HTML olarak oku
        try:
            tablolar = pdsinoptik .read_html(dosya_yolu)
            for i, df in enumerate(tablolar):
                if df.shape[1] < 5:
                    continue
                df["sayfa"] = f"gun_{i+1}"
                tum_kayitlar.append(df)
        except Exception as html_e:
            try:
                tablolar = pd.read_html(dosya_yolu, flavor='bs4')
                for i, df in enumerate(tablolar):
                    if df.shape[1] < 5:
                        continue
                    df["sayfa"] = f"gun_{i+1}"
                    tum_kayitlar.append(df)
            except Exception as html_bs4_e:
                try:
                    try:
                        df_csv = pd.read_csv(dosya_yolu, sep='\t', encoding='utf-8')
                    except UnicodeDecodeError:
                        df_csv = pd.read_csv(dosya_yolu, sep='\t', encoding='cp1254')
                    if len(df_csv.columns) < 3:
                        try:
                            df_csv = pd.read_csv(dosya_yolu, sep=None, engine='python', encoding='utf-8')
                        except UnicodeDecodeError:
                            df_csv = pd.read_csv(dosya_yolu, sep=None, engine='python', encoding='cp1254')
                    if len(df_csv.columns) >= 5:
                        df_csv["sayfa"] = "gun_1"
                        tum_kayitlar.append(df_csv)
                except Exception as csv_e:
                    raise ValueError(f"Dosyadan hiç tablo okunamadı (Excel/HTML/CSV hataları):\nExcel: {e}\nHTML: {html_e}\nHTML(bs4): {html_bs4_e}\nCSV: {csv_e}")

    if not tum_kayitlar:
        raise ValueError("Dosyadan hiç tablo okunamadı (Excel/HTML/CSV değil)")

    df_all = pd.concat(tum_kayitlar, ignore_index=True)

    # kolonları temizle
    df_all.columns = df_all.columns.astype(str).str.strip().str.lower()
    df_all = df_all.loc[:, ~df_all.columns.str.startswith("unnamed")]

    return df_all

def excel_dosyasini_oku(dosya_yolu, header_satiri):
    """
    Excel dosyasındaki TÜM sayfaları okur ve tek DataFrame döndürür
    """
    if dosya_yolu.lower().endswith(".xls"):
        engine = "xlrd"
    else:
        engine = "openpyxl"

    excel = pd.ExcelFile(dosya_yolu, engine=engine)
    tum_kayitlar = []

    for sayfa in excel.sheet_names:
        try:
            df = pd.read_excel(excel, sheet_name=sayfa, header=header_satiri)
            if df.empty:
                continue
            df["sayfa"] = sayfa
            tum_kayitlar.append(df)
        except Exception as e:
            print(f"{sayfa} okunamadı → {e}")

    if not tum_kayitlar:
        raise ValueError(f"{os.path.basename(dosya_yolu)} içinden hiç veri okunamadı")

    df_all = pd.concat(tum_kayitlar, ignore_index=True)

    # kolon isimlerini güvenli hale getir
    df_all.columns = df_all.columns.astype(str).str.strip().str.lower()
    df_all = df_all.loc[:, ~df_all.columns.str.startswith("unnamed")]

    return df_all

def sutun_adi_normalize_et(col):
    """Sütun isimlerini standart hale getirir ve varyasyonları eşleştirir."""
    c = str(col).strip().lower()
    
    # 1. Kesin ve İçerik Bazlı Temel Eşleşmeler
    if c in ['istasyon', 'ist', 'ist_no', 'ist no', 'station']: return 'istasyon_no'
    if c in ['tarih', 'date', 'gün', 'gun']: return 'sayfa'

    # GMT / Saat (Genişletilmiş Arama)
    if 'gmt' in c or 'utc' in c: return 'gmt'
    # "saat" kelimesi geçiyorsa ama "yağış", "güneş" vb. değilse (Örn: "Rasat Saati")
    if 'saat' in c and not any(x in c for x in ['güneş', 'sun', 'yağış', 'yagis', 'süresi', 'suresi', 'duration']): return 'gmt'
    if c in ['zaman', 'time', 'hour']: return 'gmt'
    
    # 2. İçerik Bazlı Eşleşmeler (Kelime arama)
    
    # Sıcaklıklar (Sıra önemli: Toprak -> Max/Min -> İşba -> Kuru)
    if 'toprak' in c or 'grass' in c or c == 'tg': return 'tg'
    if 'maks' in c or 'max' in c or c == 'tx': return 'tx'
    if 'min' in c or c == 'tn': return 'tn'
    if 'işba' in c or 'isba' in c or 'dew' in c or c == 'td': return 'td'
    if '%' in c or 'nem' in c or c == 'rh': return 'rh'
    if c == 't' or 'kuru' in c or 'dry' in c or 'temp' in c or 'sıcaklık' in c or 'sicaklik' in c: return 't'
    
    # Basınç
    if '4p' in c or 'deniz' in c or 'msl' in c or c == 'p0' or 'qff' in c or 'qnh' in c: return 'p0'
    if '3po' in c or 'aktüel' in c or 'aktuel' in c or 'istasyon basıncı' in c or 'stn' in c or 'qfe' in c or c == 'p' or 'basınç' in c or 'basinc' in c: return 'p'
    if 'değişim' in c or 'degisim' in c or c == 'ppp': return 'ppp'
    if 'karakteristik' in c or 'a' == c: return 'a'
    
    # Rüzgar
    if 'maksimum rüzgar' in c or 'max rüzgar' in c or c == 'g910': return 'g910'
    if 'hamle' in c or 'gust' in c or c == 'g911': return 'g911'
    if 'hız' in c or 'hiz' in c or 'speed' in c or c == 'ff': return 'ff'
    if 'yön' in c or 'yon' in c or 'dir' in c or c == 'dd': return 'dd'
    
    # Bulut
    if 'alçak' in c or 'alcak' in c or c == 'nh': return 'nh'
    if 'kapalılık' in c or 'total' in c or 'n' == c or 't. kp' in c: return 'n'
    if c == 'cl': return 'cl'
    if c == 'cm': return 'cm'
    if c == 'ch': return 'ch'
    if 'yükseklik' in c or 'yukseklik' in c or c == 'h' or 'dikine' in c: return 'h'
    
    # Görüş
    if c == 'vv' or 'görüş' in c or 'gorus' in c or 'vis' in c or 'hakim' in c: return 'vv'
    
    # Yağış / Hadise
    if 'yağış' in c or 'yagis' in c or c == 'rrr': return 'rrr'
    if 'süresi' in c or 'suresi' in c or c == 'tr': return 'tr'
    
    if str(col).strip() == 'WW': return 'w1' # Büyük harf WW -> w1 (Geçmiş Hava)
    if '2. grup' in c: return 'ww2'
    if '3. grup' in c: return 'ww3'
    if 'halihazır' in c or 'hali' in c or 'present' in c or c == 'ww' or '1. grup' in c: return 'ww'
    if 'geçmiş1' in c or 'gecmis1' in c or c == 'w1': return 'w1'
    if 'geçmiş2' in c or 'gecmis2' in c or c == 'w2': return 'w2'
    if 'geçmiş' in c or 'gecmis' in c: return 'w1'
    
    # İndikatörler
    if c == 'ir' or c == 'i r' or 'indikatör' in c: return 'ir' # 'ind' kelimesi 'Wind' içinde geçtiği için kaldırıldı!
    if c == 'ix' or c == 'i x': return 'ix'
    
    # 924 Grubu ve Hadise Kayıtları
    if '924' in c: return 'g924'
    if 'hadise' in c and ('kayıt' in c or 'kayit' in c or 'defter' in c): return 'hadise_kayit'

    # Personel / Gözlemci
    if any(x in c for x in ['rasatçı', 'rasatci', 'gözlemci', 'gozlemci', 'personel', 'kullanıcı', 'kullanici', 'operator', 'user', 'rstçı', 'p. no']): return 'personel'

    return c

def header_bul(df):
    safe_keywords = [
        "sıcaklık", "temp", "gmt", "saat", "zaman", "time", "utc", "ir", "i r", "ix", "i x", 
        "ww", "halihazır", "present", "hadise", "istasyon", "ist no", "station", "rüzgar", "wind", 
        "yön", "hız", "basınç", "pressure", "aktüel", "yağış", "precip", "rrr", "bulut", "cloud", 
        "kapalılık", "görüş", "vis", "vv", "3po", "4p", "qfe", "qnh", "t. kp.", "1. grup", "hakim", 
        "kuru", "işba", "bülten", "bulten", "rasat", "tür", "kayıt", "gün", "gun", "metar", "speci", 
        "mesaj", "tipi", "dikine", "rstçı", "p. no"
    ]
    exact_keywords = ["t", "p", "n", "ff", "dd", "h", "a"]

    for i in range(min(40, len(df))):
        row_values = df.iloc[i].astype(str).str.lower().values
        match_count = 0
        for val in row_values:
            v_clean = str(val).strip()
            if v_clean in exact_keywords:
                match_count += 1
            elif any(k in v_clean for k in safe_keywords):
                match_count += 1
        if match_count >= 3:
            return i
    return None

def dosya_oku_akilli(dosya_yolu):
    tum_kayitlar = []
    ekstra_veriler = {} # Tek sayılı (kapalı rasat) sayfalardaki verileri tutar
    
    try:
        # 🔹 Önce Excel gibi dene
        excel = pd.ExcelFile(dosya_yolu)
        for sayfa in excel.sheet_names:
            sayfa_str = str(sayfa).lower()
            is_odd_sheet = False
            cift_sayfa_adi = None
            
            if "sheet" in sayfa_str or "sayfa" in sayfa_str:
                rakamlar = re.findall(r'\d+', sayfa_str)
                if rakamlar:
                    sayfa_no = int(rakamlar[0])
                    if sayfa_no % 2 != 0:
                        is_odd_sheet = True
                        if sayfa_no != 1:
                            cift_sayfa_adi = sayfa_str.replace(str(sayfa_no), str(sayfa_no - 1))
            elif sayfa_str.startswith("k-"):
                is_odd_sheet = True
                cift_sayfa_adi = sayfa_str.replace("k-", "a-")
                            
            if is_odd_sheet:
                if cift_sayfa_adi:
                    try:
                        df_odd = pd.read_excel(excel, sheet_name=sayfa, header=None)
                        extracted = {}
                        # Excel Hücreleri -> (Satır İndeksi, Sütun İndeksi) | 0 Tabanlı
                        mapping = {
                            (14, 23): 'buhar', (14, 19): 'e_kar', (14, 27): 'rad_tipi',
                            (22, 14): 'top_ustu_min', (22, 19): 'g931', (25, 14): 'rrr_toplam',
                            (25, 19): 'g932', (22, 27): 'deniz_suyu', (22, 23): 'gunes',
                            (17, 23): 'buh_alet_tipi', (17, 27): 'radyasyon'
                        }
                        for (r, c), col_name in mapping.items():
                            try:
                                if r < len(df_odd) and c < len(df_odd.columns):
                                    val = df_odd.iloc[r, c]
                                    if pd.notna(val) and str(val).strip() != "":
                                        extracted[col_name] = str(val)
                            except: pass
                        orig_cift_sayfa = None
                        for s in excel.sheet_names:
                            if str(s).lower() == cift_sayfa_adi:
                                orig_cift_sayfa = s
                                break
                        if orig_cift_sayfa and extracted:
                            ekstra_veriler[orig_cift_sayfa] = extracted
                    except: pass
                continue

            ham = pd.read_excel(excel, sheet_name=sayfa, header=None)
            h = header_bul(ham)
            if h is None:
                # Fallback: Header bulunamazsa 5. satırı (indeks 4) varsay
                if len(ham) > 5: h = 3 # DÜZELTME: 5. Satır veri ise, başlık 
            df = pd.read_excel(excel, sheet_name=sayfa, header=h)
            df["sayfa"] = sayfa
            tum_kayitlar.append(df)

    except Exception:
        pass

    # Eğer Excel okuması başarısız olduysa veya başlık bulunamadıysa HTML olarak dene
    if not tum_kayitlar:
        try:
            # 🔹 Excel değilse HTML'dir
            tablolar = pd.read_html(dosya_yolu)
            for i, ham in enumerate(tablolar):
                h = header_bul(ham)
                if h is None:
                    if len(ham) > 5: h = 3
                    else: continue
                df = ham.iloc[h+1:].copy()
                df.columns = ham.iloc[h]
                df["sayfa"] = f"{i+1:02d}"
                tum_kayitlar.append(df)
        except Exception:
            try:
                tablolar = pd.read_html(dosya_yolu, flavor='bs4')
                for i, ham in enumerate(tablolar):
                    h = header_bul(ham)
                    if h is None:
                        if len(ham) > 5: h = 3
                        else: continue
                    df = ham.iloc[h+1:].copy()
                    df.columns = ham.iloc[h]
                    df["sayfa"] = f"{i+1:02d}"
                    tum_kayitlar.append(df)
            except Exception:
                try:
                    try:
                        df_csv = pd.read_csv(dosya_yolu, sep='\t', encoding='utf-8')
                    except UnicodeDecodeError:
                        df_csv = pd.read_csv(dosya_yolu, sep='\t', encoding='cp1254')
                    if len(df_csv.columns) < 3:
                        try:
                            df_csv = pd.read_csv(dosya_yolu, sep=None, engine='python', encoding='utf-8')
                        except UnicodeDecodeError:
                            df_csv = pd.read_csv(dosya_yolu, sep=None, engine='python', encoding='cp1254')
                    h = header_bul(df_csv)
                    if h is None:
                        if len(df_csv) > 5: h = 3
                        else: h = 0
                    df = df_csv.iloc[h+1:].copy()
                    df.columns = df_csv.iloc[h]
                    df["sayfa"] = "01"
                    tum_kayitlar.append(df)
                except Exception:
                    pass

    if not tum_kayitlar:
        raise ValueError(f"Dosya formatı uygun değil veya veri bulunamadı!\nDosya: {os.path.basename(dosya_yolu)}\nLütfen dosyanın bozuk olmadığından ve doğru formatta (Excel/HTML/CSV) olduğundan emin olun.")

    df_all = pd.concat(tum_kayitlar, ignore_index=True)

    # 🛠️ Kolon isimlerini normalize et (Türkçe karakter, farklı isimlendirme vb. çözümü)
    new_cols = []
    for col in df_all.columns:
        new_cols.append(sutun_adi_normalize_et(col))
    df_all.columns = new_cols
    # df_all = df_all.loc[:, ~df_all.columns.str.startswith("unnamed")] # KULLANICI İSTEĞİ: Veri kaybını önlemek için Unnamed sütunları silme
    # Aynı isimli sütunları (örn: ww ve WW -> ww, ww) teke düşür
    if any(df_all.columns.duplicated()):
        benzersiz_sutunlar = []
        for col in df_all.columns.unique():
            alt_df = df_all[[col]]
            if alt_df.shape[1] > 1:
                benzersiz_sutunlar.append(alt_df.bfill(axis=1).iloc[:, 0])
            else:
                benzersiz_sutunlar.append(alt_df.iloc[:, 0])
        df_all = pd.concat(benzersiz_sutunlar, axis=1)

    if ekstra_veriler:
        def safe_gmt(x):
            try:
                v = str(x).strip().upper().replace('Z', '')
                if ':' in v: v = v.split(':')[0]
                return float(v)
            except: return float('nan')
            
        gmt_num = df_all['gmt'].apply(safe_gmt) if 'gmt' in df_all.columns else pd.Series(dtype=float)

        for cift_sayfa, ekstra in ekstra_veriler.items():
            mask_06 = (df_all['sayfa'].astype(str).str.lower() == str(cift_sayfa).lower()) & (gmt_num == 6.0)
            mask_18 = (df_all['sayfa'].astype(str).str.lower() == str(cift_sayfa).lower()) & (gmt_num == 18.0)
            
            for k, v in ekstra.items():
                if k not in df_all.columns:
                    df_all[k] = pd.NA
                if mask_06.any():
                    df_all.loc[mask_06, k] = v
                elif mask_18.any():
                    df_all.loc[mask_18, k] = v

    return df_all

def aylik_rapor_olustur():
    # 1️⃣ DOSYA SEÇ (Ana Thread)
    try:
        messagebox.showinfo("1. Adım", "SİNOPTİK dosyasını seç")
        sin_yolu = filedialog.askopenfilename(filetypes=[("Excel", "*.xls *.xlsx")])
        if not sin_yolu:
            return

        messagebox.showinfo("2. Adım", "METAR dosyasını seç")
        metar_yolu = filedialog.askopenfilename(filetypes=[("Excel", "*.xls *.xlsx")])
        if not metar_yolu:
            return

        # 📅 TARİH BİLGİSİ AL
        # Şu anki zamana göre varsayılan değerleri ayarla (Genelde geçen ayın raporu alınır)
        simdi = datetime.datetime.now()
        varsayilan_ay = simdi.month - 1 if simdi.month > 1 else 12
        varsayilan_yil = simdi.year if simdi.month > 1 else simdi.year - 1

        yil = simpledialog.askinteger("Yıl", "Rapor Yılı:", initialvalue=varsayilan_yil, minvalue=2000, maxvalue=2030)
        if not yil: return
        ay = simpledialog.askinteger("Ay", "Rapor Ayı:", initialvalue=varsayilan_ay, minvalue=1, maxvalue=12)
        if not ay: return
    except Exception as e:
        messagebox.showerror("Hata", str(e))
        return

    # İşlem başladığında butonu pasif yap
    btn_run.config(state=tk.DISABLED, text="İşleniyor... Lütfen Bekleyin")

    def islem_yurut():
        try:
            # 2️⃣ OKU
            df_sin = dosya_oku_akilli(sin_yolu)
            df_metar = dosya_oku_akilli(metar_yolu)

            def tarih_olustur(val):
                try:
                    val_str = str(val).lower()
                    
                    match = re.search(r'\b(\d{2}\.\d{2}\.\d{4})\b', val_str)
                    if match:
                        return match.group(1)
                        
                    # Sadece rakamları al (Sheet1 -> 1, 01 -> 1)
                    gun_str = ''.join(filter(str.isdigit, val_str))
                    if not gun_str: return None
                    gun = int(gun_str)

                    # MGM Formatı: Sheet2=1.gün ... Sheet62=31.gün
                    # "sheet", "sayfa" kelimesi geçiyorsa veya sadece sayıysa (örn: "2")
                    if "sheet" in val_str or "sayfa" in val_str or val_str.strip().isdigit():
                        # Sheet2 -> 1. Gün, Sheet4 -> 2. Gün mantığı
                        # Formül: SayfaNo / 2
                        if gun >= 2 and gun % 2 == 0:
                            gun = gun // 2
                        else:
                            return None

                    return pd.Timestamp(year=yil, month=ay, day=gun).strftime('%d.%m.%Y')
                except:
                    return None

            # 📝 OKUMA RAPORU: Hangi sayfa hangi tarihe denk geldi?
            okuma_raporu = ""
            try:
                if not df_sin.empty and "sayfa" in df_sin.columns:
                    # Sayfaları numarasına göre sırala
                    sheets = sorted(df_sin["sayfa"].unique().astype(str), key=lambda x: int(''.join(filter(str.isdigit, x))) if any(c.isdigit() for c in x) else 0)
                    mapped = []
                    for s in sheets:
                        t = tarih_olustur(s)
                        if t:
                            mapped.append(f"[{s}] -> {t}")
                    
                    okuma_raporu = "SİNOPTİK SAYFA - TARİH EŞLEŞMELERİ:\n" + "-"*40 + "\n"
                    if mapped:
                        okuma_raporu += "\n".join(mapped)
                    else:
                        okuma_raporu += "⚠️ Hiçbir sayfa tarihe eşleştirilemedi! (Sayfa isimlerini kontrol edin)"
            except: pass

            for df in (df_sin, df_metar):
                if "gmt" in df.columns:
                    # "03:00", "00Z" gibi formatları temizle
                    df["gmt"] = df["gmt"].astype(str).str.upper().str.replace('Z', '').str.strip()
                    df["gmt"] = df["gmt"].str.split(':').str[0]
                    df["gmt"] = pd.to_numeric(df["gmt"], errors="coerce")
                    
                    # 🧹 TEMİZLİK: GMT değeri sayı olmayan (NaN) satırları sil (Başlık tekrarları vb.)
                    df.dropna(subset=["gmt"], inplace=True)
                
                numeric_cols = ['ir', 'ix', 'rrr', 'ww', 'w1', 'w2', 't', 'td', 'n', 'nh', 'cl', 'cm', 'ch', 'dd', 'ff', 'vv', 'a', 'ppp', 'tx', 'tn', 'tg', 'p', 'p0', 'e', 'h', 'tr', 'g924']
                for col in numeric_cols:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors="coerce")

                df["sayfa"] = df["sayfa"].astype(str)
                df["sayfa"] = df["sayfa"].apply(tarih_olustur)
                
                # Geçersiz sayfaları temizle (None dönenler)
                df.dropna(subset=["sayfa"], inplace=True)

            # GMT Sütunu Kontrolü
            if "gmt" not in df_sin.columns:
                # 1. Otomatik İçerik Kontrolü (0, 3, 6... içeren sütunu bul)
                gmt_col_found = None
                for col in df_sin.columns:
                    try:
                        vals = df_sin[col].dropna().unique()
                        # 0, 3, 6... 21 değerlerinden en az 4 tanesini içeriyorsa GMT'dir
                        target_hours = {0, 3, 6, 9, 12, 15, 18, 21, '00', '03', '06', '09', '12', '15', '18', '21'}
                        match_count = sum(1 for v in vals if v in target_hours)
                        if match_count >= 4:
                            gmt_col_found = col
                            break
                    except: pass
                
                if gmt_col_found:
                    df_sin.rename(columns={gmt_col_found: "gmt"}, inplace=True)
                # 2. Kullanıcı Bildirimi: D Sütunu (Index 3)
                elif len(df_sin.columns) > 3:
                    df_sin.rename(columns={df_sin.columns[3]: "gmt"}, inplace=True)
                # 3. Varsayılan: İlk Sütun
                elif len(df_sin.columns) > 0:
                    df_sin.rename(columns={df_sin.columns[0]: "gmt"}, inplace=True)

                if "gmt" not in df_sin.columns:
                    raise ValueError(f"Sinoptik dosyasında 'GMT' (Saat) sütunu bulunamadı!\nAlgılanan Sütunlar: {list(df_sin.columns)}")
            
            if "gmt" not in df_metar.columns:
                if len(df_metar.columns) > 0:
                    df_metar.rename(columns={df_metar.columns[0]: "gmt"}, inplace=True)
                
                # GÜVENLİK: Eksik sütunlardan kaynaklanan KeyError: 'sayfa' çökmelerini önlemek için
                if "sayfa" not in df_sin.columns: df_sin["sayfa"] = "1"
                if "gmt" not in df_sin.columns: df_sin["gmt"] = 0.0
                if "sayfa" not in df_metar.columns: df_metar["sayfa"] = "1"
                if "gmt" not in df_metar.columns: df_metar["gmt"] = 0.0

                if "gmt" not in df_metar.columns:
                    raise ValueError(f"METAR dosyasında 'GMT' (Saat) sütunu bulunamadı!\nAlgılanan Sütunlar: {list(df_metar.columns)}")

            # ️  RASATLAR Sütununu Oluştur (Sinoptik verilerinden)
            # Kullanıcı isteği: Sheet3 vb. sayfalardan D-AO sütunlarını al (Tüm veri sütunları)
            if not df_sin.empty:
                def raw_rasat_olustur(row):
                    items = []
                    parts = []
                    
                    def to_int_str(val, length=1):
                        if pd.isna(val) or str(val).strip() == "": return "/" * length
                        try:
                            return str(int(float(val))).zfill(length)
                        except: return "/" * length

                    ir = to_int_str(row.get('ir'), 1)
                    ix = to_int_str(row.get('ix'), 1)
                    h = to_int_str(row.get('h'), 1)
                    vv = to_int_str(row.get('vv'), 2)
                    parts.append(f"{ir}{ix}{h}{vv}")
                    
                    n = to_int_str(row.get('n'), 1)
                    dd = to_int_str(row.get('dd'), 2)
                    ff = to_int_str(row.get('ff'), 2)
                    parts.append(f"{n}{dd}{ff}")
                    
                    t = row.get('t')
                    if pd.notna(t) and str(t).strip() != "":
                        try:
                            t_val = float(t)
                            sn = "1" if t_val < 0 else "0"
                            t_str = str(int(round(abs(t_val) * 10))).zfill(3)
                            parts.append(f"1{sn}{t_str}")
                        except: pass
                        
                    td = row.get('td')
                    if pd.notna(td) and str(td).strip() != "":
                        try:
                            td_val = float(td)
                            sn = "1" if td_val < 0 else "0"
                            td_str = str(int(round(abs(td_val) * 10))).zfill(3)
                            parts.append(f"2{sn}{td_str}")
                        except: pass
                        
                    p = row.get('p')
                    if pd.notna(p) and str(p).strip() != "":
                        try:
                            p_val = float(p)
                            p_str = str(int(round(p_val * 10)))[-4:]
                            parts.append(f"3{p_str}")
                        except: pass
                        
                    p0 = row.get('p0')
                    if pd.notna(p0) and str(p0).strip() != "":
                        try:
                            p0_val = float(p0)
                            p0_str = str(int(round(p0_val * 10)))[-4:]
                            parts.append(f"4{p0_str}")
                        except: pass
                        
                    a = row.get('a')
                    ppp = row.get('ppp')
                    if pd.notna(a) and str(a).strip() != "" and pd.notna(ppp) and str(ppp).strip() != "":
                        try:
                            a_str = str(int(float(a)))
                            ppp_str = str(int(round(float(ppp) * 10))).zfill(3)
                            parts.append(f"5{a_str}{ppp_str}")
                        except: pass
                        
                    rrr = row.get('rrr')
                    tr = row.get('tr')
                    if pd.notna(rrr) and str(rrr).strip() != "" and pd.notna(tr) and str(tr).strip() != "":
                        try:
                            rrr_str = str(int(float(rrr))).zfill(3)
                            tr_str = str(int(float(tr)))
                            parts.append(f"6{rrr_str}{tr_str}")
                        except: pass
                        
                    ww = to_int_str(row.get('ww'), 2)
                    w1 = to_int_str(row.get('w1'), 1)
                    w2 = to_int_str(row.get('w2'), 1)
                    if ww != "//" or w1 != "/" or w2 != "/":
                        parts.append(f"7{ww}{w1}{w2}")
                        
                    nh = to_int_str(row.get('nh'), 1)
                    cl = to_int_str(row.get('cl'), 1)
                    cm = to_int_str(row.get('cm'), 1)
                    ch = to_int_str(row.get('ch'), 1)
                    if nh != "/" or cl != "/" or cm != "/" or ch != "/":
                        parts.append(f"8{nh}{cl}{cm}{ch}")
                        
                    sec333 = []
                    tx = row.get('tx')
                    if pd.notna(tx) and str(tx).strip() != "":
                        try:
                            tx_val = float(tx)
                            sn = "1" if tx_val < 0 else "0"
                            tx_str = str(int(round(abs(tx_val) * 10))).zfill(3)
                            sec333.append(f"1{sn}{tx_str}")
                        except: pass
                        
                    tn = row.get('tn')
                    if pd.notna(tn) and str(tn).strip() != "":
                        try:
                            tn_val = float(tn)
                            sn = "1" if tn_val < 0 else "0"
                            tn_str = str(int(round(abs(tn_val) * 10))).zfill(3)
                            sec333.append(f"2{sn}{tn_str}")
                        except: pass
                        
                    tg = row.get('tg')
                    e = row.get('e')
                    if pd.notna(tg) and str(tg).strip() != "" and pd.notna(e) and str(e).strip() != "":
                        try:
                            e_str = str(int(float(e)))
                            tg_val = float(tg)
                            sn = "1" if tg_val < 0 else "0"
                            tg_str = str(int(round(abs(tg_val)))).zfill(2)
                            sec333.append(f"3{e_str}{sn}{tg_str}")
                        except: pass
                        
                    g910 = row.get('g910')
                    if pd.notna(g910) and str(g910).strip() != "":
                        try:
                            g910_str = str(int(float(g910))).zfill(2)
                            sec333.append(f"910{g910_str}")
                        except: pass
                        
                    g911 = row.get('g911')
                    if pd.notna(g911) and str(g911).strip() != "":
                        try:
                            g911_str = str(int(float(g911))).zfill(2)
                            sec333.append(f"911{g911_str}")
                        except: pass
                        
                    g931 = row.get('g931')
                    if pd.notna(g931) and str(g931).strip() != "":
                        try:
                            g931_str = str(int(float(g931))).zfill(2)
                            sec333.append(f"931{g931_str}")
                        except: pass
                        
                    g932 = row.get('g932')
                    if pd.notna(g932) and str(g932).strip() != "":
                        try:
                            g932_str = str(int(float(g932))).zfill(2)
                            sec333.append(f"932{g932_str}")
                        except: pass
                    
                    g960 = row.get('g960')
                    if pd.notna(g960) and str(g960).strip() != "":
                        try:
                            g960_str = str(int(float(g960))).zfill(2)
                            sec333.append(f"960{g960_str}")
                        except: pass

                    if sec333:
                        parts.append("333")
                        parts.extend(sec333)

                    return " ".join(parts)
                df_sin["RASATLAR"] = df_sin.apply(raw_rasat_olustur, axis=1)

            #  ŞABLON OLUŞTURMA (Tüm günler ve 00-21 arası saatler)
            _, son_gun = calendar.monthrange(yil, ay)
            sablon_data = []
            for d in range(1, son_gun + 1):
                t_str = pd.Timestamp(year=yil, month=ay, day=d).strftime('%d.%m.%Y')
                for h in [0, 3, 6, 9, 12, 15, 18, 21]:
                    sablon_data.append({"sayfa": t_str, "gmt": h})
            df_sablon = pd.DataFrame(sablon_data)

            df_sin = df_sin.drop_duplicates(subset=["sayfa", "gmt"])

            # Sinoptik verisini şablona oturt (Eksik saatleri satır olarak ekle)
            df_sin = pd.merge(df_sablon, df_sin, on=["sayfa", "gmt"], how="left")
            df_metar_tekil = df_metar.drop_duplicates(subset=["sayfa", "gmt"], keep="last")

            def _zaman_don(dframe):
                return pd.to_datetime(
                    dframe["sayfa"].astype(str).str.strip() + " " +
                    dframe["gmt"].fillna(0).astype(int).astype(str).str.zfill(2) + ":00",
                    format="%d.%m.%Y %H:%M",
                    errors="coerce"
                )

            df_sin["dt"] = _zaman_don(df_sin)
            df_metar_tekil["dt"] = _zaman_don(df_metar_tekil)
            df_sin.sort_values("dt", inplace=True)
            df_metar_tekil.sort_values("dt", inplace=True)

            birlesik = pd.merge_asof(
                df_sin,
                df_metar_tekil,
                on="dt",
                direction="nearest",
                tolerance=pd.Timedelta(hours=3),
                suffixes=("_sin", "_metar")
            )
            birlesik.drop(columns=[c for c in ["sayfa_metar", "gmt_metar"] if c in birlesik.columns], inplace=True, errors="ignore")
            birlesik.drop(columns=["dt"], inplace=True, errors="ignore")

            if birlesik.empty:
                raise ValueError("Birleştirme sonucu boş")

            hata_dict = kurallar.HATA_SOZLUGU

            # 5️⃣ HATA ANALİZİ
            def hata_ara(s):
                kodlar, aciklamalar = [], []

                ir = s.get("ir")
                rrr = s.get("rrr")
                ix = s.get("ix")
                ww = s.get("ww")
                w1 = s.get("w1")
                w2 = s.get("w2")
                t = s.get("t")
                td = s.get("td")
                tx = s.get("tx")
                tn = s.get("tn")
                tg = s.get("tg")
                n = s.get("n")
                nh = s.get("nh")
                cl = s.get("cl")
                cm = s.get("cm")
                ch = s.get("ch")
                dd = s.get("dd")
                ff = s.get("ff")
                vv = s.get("vv")
                a = s.get("a")
                ppp = s.get("ppp")
                p = s.get("p")
                p0 = s.get("p0")
                e = s.get("e")
                h = s.get("h")
                tr = s.get("tr")
                gmt = s.get("gmt")
                g924 = s.get("g924")
                hadise_kayit = s.get("hadise_kayit")

                # Güvenli tip dönüşümü
                try: ir = int(ir) if pd.notna(ir) else None
                except: pass
                try: ix = int(ix) if pd.notna(ix) else None
                except: pass

                # 🛑 BOŞ SATIR / GEÇERSİZ SATIR KONTROLÜ
                # GMT yoksa veya kritik verilerin çoğu yoksa atla (Excel'in boş satırlarını okumasını engelle)
                if pd.isna(gmt) or (pd.isna(t) and pd.isna(td) and pd.isna(p) and pd.isna(ir)):
                    # Eğer şablondan gelen bir satırsa (GMT var ama veri yok)
                    if pd.notna(gmt):
                        return ("Veri Yok", "Veri Yok", "Bu saatte rasat verisi bulunamadı.")
                    return ("Hata Yok", "", "")

                if ir == 1 and pd.isna(rrr):
                    msg = hata_dict.get("h4", "İr=1 olduğu halde yağış grubu kodlanmamış")
                    
                    # 🔍 GEÇMİŞ METAR VE HADİSE KONTROLÜ
                    yagis_var_mi = False
                    kanitlar = []

                    # 1. Sinoptik Geçmiş Hava (W1, W2) Kontrolü
                    # 5,6,7,8,9 kodları yağış ifade eder
                    try:
                        if (pd.notna(w1) and int(w1) >= 5) or (pd.notna(w2) and int(w2) >= 5):
                            yagis_var_mi = True
                            kanitlar.append("Sinoptik W1/W2")
                    except: pass

                    # 2. METAR Kontrolü (Son 6 saat)
                    try:
                        tarih_str = s.get("sayfa") # dd.mm.yyyy
                        saat = gmt
                        gun_val, ay_val, yil_val = map(int, tarih_str.split('.'))
                        dt_suan = datetime.datetime(yil_val, ay_val, gun_val, int(saat))
                        
                        ww_col = "ww" if "ww" in df_metar.columns else None
                            
                        if ww_col:
                            metar_kesit = pd.DataFrame()
                            
                            # Saatlere göre tarama aralığı belirle
                            lookback = 3 # Varsayılan (Ara rasatlar)
                            if saat == 0: lookback = 6
                            elif saat == 12: lookback = 6
                            elif saat == 18: lookback = 12
                            elif saat == 6: lookback = 24
                            
                            dt_baslangic = dt_suan - datetime.timedelta(hours=lookback)
                            
                            if dt_baslangic.date() == dt_suan.date():
                                # Aynı gün içinde
                                mask = (df_metar["sayfa"] == tarih_str) & \
                                       (df_metar["gmt"] > dt_baslangic.hour) & \
                                       (df_metar["gmt"] <= saat)
                                metar_kesit = df_metar[mask]
                            else:
                                # Önceki güne sarkıyor
                                prev_day_str = dt_baslangic.strftime('%d.%m.%Y')
                                
                                # Önceki gün: Başlangıç saatinden gün sonuna kadar
                                mask_prev = (df_metar["sayfa"] == prev_day_str) & (df_metar["gmt"] >= dt_baslangic.hour)
                                # Bugün: 00'dan şu anki saate kadar
                                mask_curr = (df_metar["sayfa"] == tarih_str) & (df_metar["gmt"] <= saat)
                                
                                metar_kesit = pd.concat([df_metar[mask_prev], df_metar[mask_curr]])
                            
                            for w in metar_kesit[ww_col]:
                                try:
                                    w_int = int(w)
                                    # 20-27: Geçmiş Yağış, 29: Geçmiş Oraj, 50-99: Yağış
                                    # (10 Pus, 28 Sis, 40-49 Sis HARİÇ)
                                    if (20 <= w_int <= 27) or (w_int == 29) or (50 <= w_int <= 99):
                                        yagis_var_mi = True
                                        kanitlar.append(f"METAR ww={w_int}")
                                        break
                                except:
                                    # Sayısal değilse Metin (String) kontrolü yap (BCFG, FG, BR vb. için)
                                    w_str = str(w).upper()
                                    # Sadece kesin yağış kodlarını ara (FG, BR, BCFG, HZ vb. yağış değildir)
                                    # RA:Yağmur, SN:Kar, DZ:Çisenti, GR:Dolu, GS:Küçük Dolu, SG:Kar Taneleri, PL:Buz Taneleri
                                    yagis_kodlari = ["RA", "SN", "DZ", "GR", "GS", "SG", "PL", "UP"]
                                    if any(kod in w_str for kod in yagis_kodlari):
                                        yagis_var_mi = True
                                        kanitlar.append(f"METAR ww={w_str}")
                                        break
                    except Exception: pass

                    if yagis_var_mi:
                        msg += f" (TEYİTLİ: {', '.join(kanitlar)} var -> RRR girilmeli)"
                        kodlar.append("h4 (TEYİTLİ)")
                    else:
                        msg += " (DİKKAT: Geçmişte yağış kaydı/METAR bulunamadı, İr=1 hatalı olabilir)"
                        kodlar.append("h4 (DİKKAT)")
                    aciklamalar.append(msg)

                # h5: İr=3 olduğu halde yağış grubu kodlanmış
                if ir == 3 and pd.notna(rrr):
                    kodlar.append("h5")
                    aciklamalar.append(hata_dict.get("h5"))

                # h11: İx=1 iken hadise grubu verilmemiş
                if ix == 1 and pd.isna(ww):
                    kodlar.append("h11")
                    aciklamalar.append(hata_dict.get("h11"))
                # h12: İx=2 iken hadise grubu verilmiş
                elif ix == 2 and pd.notna(ww):
                    kodlar.append("h12")
                    aciklamalar.append(hata_dict.get("h12"))

                # h13: İx 1 ve 2 den farklı kodlanmış
                if pd.notna(ix) and ix not in [1, 2, 3]: # 3 bazen kullanılabiliyor ama genelde 1-2
                    kodlar.append("h13")
                    aciklamalar.append(hata_dict.get("h13"))

                # --- EKSTRA KONTROLLER (h1...h262 Kapsamı) ---
                
                # Nisbi Nem Hesabı (Görüş/Hadise kontrolleri için)
                rh = None
                if pd.notna(t) and pd.notna(td):
                    try:
                        es = 6.112 * math.exp((17.67 * t) / (t + 243.5))
                        e_vap = 6.112 * math.exp((17.67 * td) / (td + 243.5))
                        rh = (e_vap / es) * 100
                    except: pass

                # h3: İr grubu hatalı (0-4 arası değil)
                if pd.notna(ir) and ir not in [0, 1, 2, 3, 4]:
                    kodlar.append("h3")
                    aciklamalar.append(hata_dict.get("h3"))

                # h8: İndikatör hatalı (İr veya İx limit dışı)
                if (pd.notna(ir) and ir not in [0,1,2,3,4]) or (pd.notna(ix) and ix not in [1,2,3]):
                    kodlar.append("h8")
                    aciklamalar.append(hata_dict.get("h8"))

                # h6, h75: Ara rasatta yağış grubu kodlanmış (03, 09, 15, 21)
                if gmt in [3, 9, 15, 21] and pd.notna(rrr):
                    kodlar.append("h6")
                    aciklamalar.append(hata_dict.get("h6"))
                
                # h17: h hatalı (0-9 arası değil)
                if pd.notna(h) and not (0 <= h <= 9):
                    kodlar.append("h17")
                    aciklamalar.append(hata_dict.get("h17"))

                # h22: Görüş kodu hatalı (0-99 değil)
                if pd.notna(vv) and not (0 <= vv <= 99):
                    kodlar.append("h22")
                    aciklamalar.append(hata_dict.get("h22"))

                # h26: N=0 iken bulut grubu verilmiş
                if n == 0 and ((pd.notna(nh) and nh > 0) or (pd.notna(cl) and cl > 0) or (pd.notna(cm) and cm > 0) or (pd.notna(ch) and ch > 0)):
                    kodlar.append("h26")
                    aciklamalar.append(hata_dict.get("h26"))

                # h27: Bulut grubu yok ama N > 0
                if pd.isna(nh) and pd.isna(cl) and pd.isna(cm) and pd.isna(ch) and pd.notna(n) and n > 0:
                    kodlar.append("h27")
                    aciklamalar.append(hata_dict.get("h27"))

                # h30: h=/ (NaN) iken N!=9 (CL/CM varsa h olmalı)
                if pd.isna(h) and pd.notna(n) and n != 9:
                    if (pd.notna(cl) and cl > 0) or (pd.notna(cm) and cm > 0):
                         kodlar.append("h30")
                         aciklamalar.append(hata_dict.get("h30"))

                # h31: h var iken N=9
                if pd.notna(h) and n == 9:
                    kodlar.append("h31")
                    aciklamalar.append(hata_dict.get("h31"))

                # h32: N=9 iken Ns=9 kodlanmamış (Genelde Nh kullanılır)
                if n == 9 and (pd.isna(nh) or nh != 9):
                    kodlar.append("h32")
                    aciklamalar.append(hata_dict.get("h32"))

                # h33: N=9 iken C=/ kodlanmamış (CL, CM, CH boş olmalı veya /)
                if n == 9 and (pd.notna(cl) or pd.notna(cm) or pd.notna(ch)):
                    kodlar.append("h33")
                    aciklamalar.append(hata_dict.get("h33"))

                # h34: Toplam kapalılık ile bulut kapalılığı uyumsuz (N < Nh)
                if pd.notna(n) and pd.notna(nh) and n < nh:
                    kodlar.append("h34")
                    aciklamalar.append(hata_dict.get("h34"))

                # h120: Nh > N (h34 ile benzer ama kural listesinde var)
                if pd.notna(nh) and pd.notna(n) and nh > n:
                    kodlar.append("h120")
                    aciklamalar.append(hata_dict.get("h120"))

                # h37: dd kıstas dışı (0-36 veya 99)
                if pd.notna(dd) and not ((0 <= dd <= 36) or dd == 99):
                    kodlar.append("h37")
                    aciklamalar.append(hata_dict.get("h37"))

                # h38: dd=00 iken ff!=00
                if dd == 0 and pd.notna(ff) and ff > 0:
                    kodlar.append("h38")
                    aciklamalar.append(hata_dict.get("h38"))

                # h39: dd!=00 iken ff=00
                if pd.notna(dd) and dd != 0 and pd.notna(ff) and ff == 0:
                    kodlar.append("h39")
                    aciklamalar.append(hata_dict.get("h39"))

                # h43: ff=00 iken dd yön verilmiş (h39 ile benzer mantık)
                if ff == 0 and pd.notna(dd) and dd > 0:
                    kodlar.append("h43")
                    aciklamalar.append(hata_dict.get("h43"))

                # h44: Rüzgar hızı limit dışı (>100 kt)
                if pd.notna(ff) and ff > 100:
                    kodlar.append("h44")
                    aciklamalar.append(hata_dict.get("h44"))

                # h48, h53, h187, h193: Sıcaklık Limitleri (-50 ile +60 arası)
                for val, code in [(t, "h48"), (td, "h53"), (tx, "h187"), (tn, "h193")]:
                    if pd.notna(val) and (val < -50 or val > 60):
                        kodlar.append(code)
                        aciklamalar.append(hata_dict.get(code))

                # h55, h58: Basınç Limitleri (800-1100 hPa)
                for val, code in [(p, "h55"), (p0, "h58")]:
                    if pd.notna(val) and (val < 800 or val > 1100):
                        kodlar.append(code)
                        aciklamalar.append(hata_dict.get(code))

                # h61: Tandans karakteristiği 9 olamaz
                if a == 9:
                    kodlar.append("h61")
                    aciklamalar.append(hata_dict.get("h61"))

                # h63: Tandans karakteri 4 iken değişim 0 olur
                if a == 4 and pd.notna(ppp) and ppp != 0:
                    kodlar.append("h63")
                    aciklamalar.append(hata_dict.get("h63"))

                # h62: a!=0,4,5 iken ppp=0
                if pd.notna(a) and a not in [0, 4, 5] and pd.notna(ppp) and ppp == 0:
                    kodlar.append("h62")
                    aciklamalar.append(hata_dict.get("h62"))

                # h64: ppp çok fazla (>20 hPa)
                if pd.notna(ppp) and ppp > 20:
                    kodlar.append("h64")
                    aciklamalar.append(hata_dict.get("h64"))

                # h70: Yağış miktarı 000 kodlanmaz
                if rrr == 0:
                    kodlar.append("h70")
                    aciklamalar.append(hata_dict.get("h70"))

                # h71: RRR çok fazla (>400 mm)
                if pd.notna(rrr) and rrr > 400:
                    kodlar.append("h71")
                    aciklamalar.append(hata_dict.get("h71"))

                # h74: tR kontrolü (0-9 arası)
                if pd.notna(tr) and not (0 <= tr <= 9):
                    kodlar.append("h74")
                    aciklamalar.append(hata_dict.get("h74"))

                # h76: Hadise yok (ww<4) ama RRR var
                if pd.notna(ww) and ww < 4 and pd.notna(rrr) and rrr > 0:
                    kodlar.append("h76")
                    aciklamalar.append(hata_dict.get("h76"))

                # h78: Halihazır hava 00 olamaz
                if ww == 0:
                    kodlar.append("h78")
                    aciklamalar.append(hata_dict.get("h78"))

                # h79: Görüş 10 Km.den az (kod < 60) iken hadise verilmemiş
                if pd.notna(vv) and vv < 60 and (pd.isna(ww) or ww < 4):
                    kodlar.append("h79")
                    aciklamalar.append(hata_dict.get("h79"))

                # h80: ww=01-03 iken W1>2 olamaz
                if pd.notna(ww) and 1 <= ww <= 3 and pd.notna(w1) and w1 > 2:
                     kodlar.append("h80")
                     aciklamalar.append(hata_dict.get("h80"))

                # h81, h82, h84: Görüş ve Nem Kontrolleri
                if ww == 4 and ((pd.notna(vv) and vv >= 90) or (rh is not None and rh >= 70)):
                    kodlar.append("h81")
                    aciklamalar.append(hata_dict.get("h81"))
                if ww == 5 and ((pd.notna(vv) and vv >= 90) or (rh is not None and rh >= 70)):
                    kodlar.append("h82")
                    aciklamalar.append(hata_dict.get("h82"))
                if ww == 10 and ((pd.notna(vv) and (vv < 10 or vv >= 90)) or (rh is not None and rh < 70)):
                    kodlar.append("h84")
                    aciklamalar.append(hata_dict.get("h84"))

                # h86: ww=17 (Oraj) -> Cb olmalı (CL=3,9)
                if ww == 17 and pd.notna(cl) and cl not in [3, 9]:
                    kodlar.append("h86")
                    aciklamalar.append(hata_dict.get("h86"))

                # h111: Yağış var (ww>=50) ama N=0
                if pd.notna(ww) and ww >= 50 and n == 0:
                    kodlar.append("h111")
                    aciklamalar.append(hata_dict.get("h111"))

                # h116: İkinci geçmiş hadise birinci geçmiş hadiseden büyük olamaz
                if pd.notna(w1) and pd.notna(w2) and w2 > w1:
                    kodlar.append("h116")
                    aciklamalar.append(hata_dict.get("h116"))

                # h118: ww<4 iken W1, W2 gerek yok
                if pd.notna(ww) and ww < 4 and (pd.notna(w1) or pd.notna(w2)):
                    kodlar.append("h118")
                    aciklamalar.append(hata_dict.get("h118"))

                # h196: Yerin hali ve toprak sıcaklık grubu koda dahil edilmedi (06 GMT)
                if gmt == 6 and pd.isna(e) and pd.isna(tg):
                    kodlar.append("h196")
                    aciklamalar.append(hata_dict.get("h196"))

                # h197: t < 0 iken E=1,2 (ıslak) olmaz
                if pd.notna(t) and t < 0 and pd.notna(e) and e in [1, 2]:
                    kodlar.append("h197")
                    aciklamalar.append(hata_dict.get("h197"))

                # --- EKSİK GRUP KONTROLLERİ (h252-h260) ---
                # Sadece en az bir verinin olduğu satırlarda eksik grup kontrolü yap (Hayalet satırları engelle)
                if not (pd.isna(t) and pd.isna(td) and pd.isna(p) and pd.isna(n)):
                    if pd.isna(t): kodlar.append("h252"); aciklamalar.append(hata_dict.get("h252"))
                    if pd.isna(td): kodlar.append("h253"); aciklamalar.append(hata_dict.get("h253"))
                    if pd.isna(p): kodlar.append("h254"); aciklamalar.append(hata_dict.get("h254"))
                    if pd.isna(p0): kodlar.append("h255"); aciklamalar.append(hata_dict.get("h255"))
                    if pd.isna(a): kodlar.append("h256"); aciklamalar.append(hata_dict.get("h256"))
                    if pd.isna(n): kodlar.append("h257"); aciklamalar.append(hata_dict.get("h257"))
                    
                    if gmt == 18 and pd.isna(tx):
                        kodlar.append("h258")
                        aciklamalar.append(hata_dict.get("h258"))
                    if gmt == 6 and pd.isna(tn):
                        kodlar.append("h259")
                        aciklamalar.append(hata_dict.get("h259"))
                    if gmt == 6 and pd.isna(tg) and pd.isna(e):
                        kodlar.append("h260")
                        aciklamalar.append(hata_dict.get("h260"))
                
                # h261: 924 grubu koda dahil edilmedi (Sütun varsa ve boşsa)
                if "g924" in s and pd.isna(g924):
                    kodlar.append("h261")
                    aciklamalar.append(hata_dict.get("h261"))

                # h262: Hadise Kayıtlarında eksiklik
                if "hadise_kayit" in s and pd.isna(hadise_kayit) and pd.notna(ww) and ww >= 4:
                    kodlar.append("h262")
                    aciklamalar.append(hata_dict.get("h262"))

                if kodlar:
                    return ("Hata Var", ", ".join(kodlar), " | ".join(aciklamalar))
                else:
                    return ("Hata Yok", "", "")

            # apply yerine list comprehension
            kayitlar = birlesik.to_dict('records')
            sonuclar = [hata_ara(satir) for satir in kayitlar]
            sonuc = pd.DataFrame(sonuclar, columns=["ANALİZ_SONUCU", "HATA_KODLARI", "HATA_ACIKLAMALARI"], index=birlesik.index)

            birlesik = pd.concat([birlesik, sonuc], axis=1)

            # Tarihe göre sırala
            try:
                birlesik["_sort_date"] = pd.to_datetime(birlesik["sayfa"], format="%d.%m.%Y", errors='coerce')
                birlesik = birlesik.sort_values(by=["_sort_date", "gmt"])
                birlesik = birlesik.drop(columns=["_sort_date"])
            except: pass

            col_map = {
                'sayfa': 'Tarih', 'gmt': 'Saat (GMT)', 'ir': 'İndikatör (ir)', 'ix': 'İndikatör (ix)',
                'h': 'Bulut Yük. (h)', 'vv': 'Görüş (VV)', 'n': 'Toplam Bulut (N)',
                'dd': 'Rüzgar Yönü (dd)', 'ff': 'Rüzgar Hızı (ff)', 't': 'Sıcaklık (T)',
                'td': 'İşba (Td)', 'p0': 'Deniz Basıncı (P0)', 'p': 'İstasyon Basıncı (P)',
                'a': 'Basınç Karakteri (a)', 'ppp': 'Basınç Değişimi (ppp)',
                'ww': 'Halihazır Hava (ww)', 'WW': 'Geçmiş Hava 1 (W1)', 'w2': 'Geçmiş Hava 2 (W2)',
                'nh': 'Alçak/Orta Bulut (Nh)', 'cl': 'Alçak Bulut (CL)', 'cm': 'Orta Bulut (CM)',
                'ch': 'Yüksek Bulut (CH)', 'tx': 'Maks. Sıcaklık (Tx)', 'tn': 'Min. Sıcaklık (Tn)',
                'tg': 'Toprak Sıcaklığı (Tg)', 'e': 'Yerin Hali (E)', 'rrr': 'Yağış Miktarı (RRR)',
                'tr': 'Yağış Süresi (tR)', 'ANALİZ_SONUCU': 'DURUM',
                'buhar': 'Buharlaşma', 'rad_tipi': 'Radyasyon Tipi', 'radyasyon': 'Radyasyon Miktarı',
                'gunes': 'Güneşlenme Süresi', 'deniz_suyu': 'Deniz Suyu Sıc.', 'rrr_toplam': 'Toplam Yağış',
                'buh_alet_tipi': 'Buhar Aleti Tipi', 'e_kar': 'Yerin Hali (Kar)',
                'top_ustu_min': 'Toprak Üstü Min.',
                'HATA_KODLARI': 'HATA KODU', 'HATA_ACIKLAMALARI': 'AÇIKLAMA',
                'RASATLAR': 'RASATLAR',
                'g924': '924 Grubu', 'hadise_kayit': 'Hadise Kayıtları',
                'Tarih': 'Tarih',
                'personel': 'Personel'
            }

            new_columns = {}

            for c in birlesik.columns:
                # _sin ve _metar eklerini koruyarak isimlendir
                base = c.replace('_sin', '').replace('_metar', '')
                suffix = '_sin' if '_sin' in c else ('_metar' if '_metar' in c else '')
                
                if base in col_map:
                    new_name = col_map[base]
                    if suffix == '_sin': new_name = f"SİNOPTİK - {new_name}"
                    elif suffix == '_metar': new_name = f"METAR - {new_name}"
                    new_columns[c] = new_name
                else:
                    new_columns[c] = c.upper()

            birlesik.rename(columns=new_columns, inplace=True)
            
            # Aynı isme sahip sütunlar oluşursa Pandas'ın çökmesini engellemek için isimleri tekilleştir
            if any(birlesik.columns.duplicated()):
                cols = pd.Series(birlesik.columns)
                for dup in cols[cols.duplicated()].unique():
                    dup_indices = cols[cols == dup].index.tolist()
                    for idx_num, idx in enumerate(dup_indices):
                        if idx_num != 0:
                            cols[idx] = f"{dup}_{idx_num}"
                birlesik.columns = cols

            # Önemli sütunları başa al
            istenen_sutunlar = ['Tarih', 'Saat (GMT)', 'DURUM', 'HATA KODU', 'RASATLAR', 'AÇIKLAMA']
            birlesik = birlesik[[c for c in istenen_sutunlar if c in birlesik.columns]]
            oncelikli = ['Tarih', 'Saat (GMT)', 'DURUM', 'HATA KODU', 'AÇIKLAMA']
            mevcut_oncelikli = [c for c in oncelikli if c in birlesik.columns]
            digerleri = [c for c in birlesik.columns if c not in mevcut_oncelikli]
            birlesik = birlesik[mevcut_oncelikli + digerleri]

            # 6️⃣ KAYDET
            masaustu = os.path.join(os.environ["USERPROFILE"], "Desktop")
            
            # Dosya ismini belirle (Eğer dosya açıksa üzerine yazamaz, yeni isim türet)
            base_filename = f"DENETIM_{yil}_{ay:02d}"
            cikti = os.path.join(masaustu, f"{base_filename}.xlsx")
            
            counter = 1
            while True:
                try:
                    with open(cikti, 'a'): pass
                    break
                except PermissionError:
                    cikti = os.path.join(masaustu, f"{base_filename} ({counter}).xlsx")
                    counter += 1
            
            if "DURUM" in birlesik.columns:
                durum_ser = birlesik["DURUM"].fillna("").astype(str).str.strip()
                hatalar_df = birlesik[~durum_ser.isin(["Hata Yok", "Ara Rasat"])]
                hata_kodu_ser = birlesik["HATA KODU"].fillna("").astype(str).str.strip() if "HATA KODU" in birlesik.columns else pd.Series([""] * len(birlesik), index=birlesik.index)
                aciklama_ser = birlesik["AÇIKLAMA"].fillna("").astype(str).str.strip() if "AÇIKLAMA" in birlesik.columns else pd.Series([""] * len(birlesik), index=birlesik.index)
                ekstra = birlesik[(durum_ser == "") & ((hata_kodu_ser != "") | (aciklama_ser != ""))]
                if not ekstra.empty:
                    hatalar_df = pd.concat([hatalar_df, ekstra]).drop_duplicates()
            else:
                hata_kodu_ser = birlesik["HATA KODU"].fillna("").astype(str).str.strip() if "HATA KODU" in birlesik.columns else pd.Series([""] * len(birlesik), index=birlesik.index)
                aciklama_ser = birlesik["AÇIKLAMA"].fillna("").astype(str).str.strip() if "AÇIKLAMA" in birlesik.columns else pd.Series([""] * len(birlesik), index=birlesik.index)
                hatalar_df = birlesik[(hata_kodu_ser != "") | (aciklama_ser != "")]

            # --- İSTATİSTİKLER ---
            from collections import Counter
            
            # 1. Hata Kodları
            tum_kodlar = []
            if not hatalar_df.empty:
                for k in hatalar_df["HATA KODU"].dropna():
                    tum_kodlar.extend([x.strip() for x in str(k).split(",") if x.strip()])
            kod_sayilari = Counter(tum_kodlar)
            df_kod_ist = pd.DataFrame(kod_sayilari.most_common(), columns=["Hata Kodu", "Adet"])

            # 2. Günlük Hata
            if "Tarih" in hatalar_df.columns:
                df_gun_ist = hatalar_df["Tarih"].value_counts().reset_index()
                df_gun_ist.columns = ["Tarih", "Hata Sayısı"]
                df_gun_ist = df_gun_ist.sort_values(by="Tarih")
            else:
                df_gun_ist = pd.DataFrame()

            # 3. Personel Hata İstatistikleri
            df_personel_ist = pd.DataFrame()
            personel_cols = [c for c in hatalar_df.columns if "Personel" in c]
            if personel_cols:
                tum_personel = []
                for col in personel_cols:
                    tum_personel.extend(hatalar_df[col].dropna().astype(str).tolist())
                
                if tum_personel:
                    p_counts = Counter(tum_personel)
                    df_personel_ist = pd.DataFrame(p_counts.most_common(), columns=["Personel", "Hata Sayısı"])

            # 3. Genel
            df_genel = pd.DataFrame([
                {"Bilgi": "Rapor Dönemi", "Değer": f"{ay}/{yil}"},
                {"Bilgi": "Toplam Kayıt", "Değer": len(birlesik)},
                {"Bilgi": "Hatalı Kayıt", "Değer": len(hatalar_df)},
                {"Bilgi": "Hata Oranı", "Değer": f"%{(len(hatalar_df)/len(birlesik)*100):.1f}" if len(birlesik) > 0 else "0"},
            ])

            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import BarChart, Reference
            from openpyxl.styles import Alignment

            with pd.ExcelWriter(cikti, engine="openpyxl") as writer:
                # ÖZET Sayfası
                df_genel.to_excel(writer, sheet_name="OZET", index=False, startrow=1, startcol=1)
                df_kod_ist.to_excel(writer, sheet_name="OZET", index=False, startrow=1, startcol=4)
                df_gun_ist.to_excel(writer, sheet_name="OZET", index=False, startrow=1, startcol=7)
                if not df_personel_ist.empty:
                    df_personel_ist.to_excel(writer, sheet_name="OZET", index=False, startrow=1, startcol=10)

                ws_ozet = writer.sheets["OZET"]
                ws_ozet["B1"] = "GENEL DURUM"
                ws_ozet["E1"] = "HATA TÜRÜ DAĞILIMI"
                ws_ozet["H1"] = "GÜNLÜK HATA DAĞILIMI"
                ws_ozet["K1"] = "PERSONEL HATA DAĞILIMI"
                for cell in ["B1", "E1", "H1", "K1"]:
                    ws_ozet[cell].font = Font(bold=True, size=12)

                # 📊 GRAFİK EKLEME (Hata Türleri)
                if not df_kod_ist.empty:
                    chart = BarChart()
                    chart.type = "col"
                    chart.style = 10
                    chart.title = "Hata Türü Dağılımı"
                    chart.y_axis.title = "Adet"
                    chart.x_axis.title = "Hata Kodu"
                    chart.legend = None

                    data = Reference(ws_ozet, min_col=6, min_row=2, max_row=2+len(df_kod_ist), max_col=6)
                    cats = Reference(ws_ozet, min_col=5, min_row=3, max_row=2+len(df_kod_ist))
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                    ws_ozet.add_chart(chart, "K2")

                birlesik.to_excel(writer, sheet_name="Detaylı_Rapor", index=False)
                birlesik.to_excel(writer, sheet_name="Detaylı_Rapor", index=False, startrow=2)
                
                if not hatalar_df.empty:
                    hatalar_df.to_excel(writer, sheet_name="SADECE_HATALAR", index=False)
                    hatalar_df.to_excel(writer, sheet_name="SADECE_HATALAR", index=False, startrow=2)

                    # 🔹 HATA KODUNA GÖRE AYRI SAYFALAR
                    tum_hatalar = set()
                    for kodlar in hatalar_df["HATA KODU"]:
                        if pd.isna(kodlar): continue
                        for kod in str(kodlar).split(","):
                            tum_hatalar.add(kod.strip())

                    for hata in sorted(tum_hatalar):
                        def hata_var_mi(val):
                            if pd.isna(val): return False
                            kodlar = [k.strip() for k in str(val).split(",")]
                            return hata in kodlar

                        df_ozel = hatalar_df[hatalar_df["HATA KODU"].apply(hata_var_mi)]
                        
                        if not df_ozel.empty:
                            sheet_name = re.sub(r'[\\/*?:\[\]]', '', str(hata))[:31]
                            try:
                                df_ozel.to_excel(writer, sheet_name=sheet_name, index=False)
                                df_ozel.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
                            except: pass

                # Başlık Stili (Kalın Font, Yeşil Arka Plan, Beyaz Yazı)
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
                
                # Hata Satırı Stili (Kırmızı Arka Plan)
                error_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                # Veri Yok Stili (Gri Arka Plan)
                missing_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

                # 📐 TÜM SAYFALAR İÇİN OTOMATİK GENİŞLİK VE STİL AYARI
                for sheet_name in writer.sheets:
                    ws = writer.sheets[sheet_name]
                    
                    # AutoFilter Ekle
                    if ws.dimensions:
                        ws.auto_filter.ref = ws.dimensions
                    if sheet_name == "OZET":
                        continue

                    # LEJANT EKLEME
                    ws["A1"] = "RENK LEJANTI:"
                    ws["A1"].font = Font(bold=True)
                    ws["B1"] = "Hatalı Değer"
                    ws["B1"].fill = error_fill
                    ws["B1"].alignment = Alignment(horizontal='center')
                    ws["C1"] = "Veri Yok"
                    ws["C1"].fill = missing_fill
                    ws["C1"].alignment = Alignment(horizontal='center')

                    # AutoFilter Ekle (Veri Başlığı 3. Satırda)
                    if ws.max_row >= 3:
                        ws.auto_filter.ref = f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"

                    durum_col_idx = None
                    hata_kodu_col_idx = None

                    # 3. Satırı (Başlık) Boya ve DURUM sütununu bul
                    for cell in ws[3]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        if cell.value and str(cell.value) == "DURUM":
                            durum_col_idx = cell.column
                        if cell.value and str(cell.value) == "HATA KODU":
                            hata_kodu_col_idx = cell.column

                    # Hata Kodu -> Sütun Adı Eşleştirmesi
                    hata_sutun_map = {
                        "h48": "Sıcaklık (T)", "h53": "İşba (Td)", "h187": "Maks. Sıcaklık (Tx)", "h193": "Min. Sıcaklık (Tn)",
                        "h55": "İstasyon Basıncı (P)", "h58": "Deniz Basıncı (P0)", "h64": "Basınç Değişimi (ppp)",
                        "h44": "Rüzgar Hızı (ff)", "h37": "Rüzgar Yönü (dd)", "h38": "Rüzgar", "h39": "Rüzgar", "h43": "Rüzgar",
                        "h70": "Yağış Miktarı (RRR)", "h71": "Yağış Miktarı (RRR)", "h74": "Yağış Süresi (tR)", "h4": "Yağış Miktarı (RRR)",
                        "h22": "Görüş (VV)", "h17": "Bulut Yük. (h)",
                        "h26": "Toplam Bulut (N)", "h27": "Toplam Bulut (N)", "h34": "Toplam Bulut (N)", "h120": "Toplam Bulut (N)",
                        "h3": "İndikatör (ir)", "h5": "İndikatör (ir)", "h8": "İndikatör",
                        "h11": "İndikatör (ix)", "h12": "İndikatör (ix)", "h13": "İndikatör (ix)",
                        "h252": "Sıcaklık (T)", "h253": "İşba (Td)", "h254": "İstasyon Basıncı (P)",
                        "h255": "Deniz Basıncı (P0)", "h257": "Toplam Bulut (N)", "h258": "Maks. Sıcaklık (Tx)", "h259": "Min. Sıcaklık (Tn)"
                    }

                    # Başlık satırını al (Sütun isminden indekse erişmek için)
                    header_cells = [str(c.value).upper() if c.value else "" for c in ws[3]]

                    # Satırları boya
                    if durum_col_idx and hata_kodu_col_idx:
                        for row in ws.iter_rows(min_row=4):
                            durum_val = row[durum_col_idx-1].value
                            hata_val = str(row[hata_kodu_col_idx-1].value) if row[hata_kodu_col_idx-1].value else ""

                            if durum_val == "Veri Yok":
                                for cell in row:
                                    cell.fill = missing_fill
                            elif durum_val != "Hata Yok":
                                # Sadece Hata Kodu hücresini boya
                                row[hata_kodu_col_idx-1].fill = error_fill
                                
                                # Hata kodlarını ayrıştır ve ilgili hücreleri boya
                                kodlar = [k.strip().split()[0] for k in hata_val.split(",")] # "h4 (TEYİTLİ)" -> "h4"
                                for kod in kodlar:
                                    target = hata_sutun_map.get(kod)
                                    if target:
                                        # Başlıkta bu kelimeyi içeren sütunları bul
                                        for idx, header in enumerate(header_cells):
                                            if target.upper() in header:
                                                row[idx].fill = error_fill

                    for column_cells in ws.columns:
                        max_length = 0
                        col_letter = get_column_letter(column_cells[0].column)
                        
                        # Kolon başlığını kontrol et (Açıklama sütunları için)
                        header_val = str(column_cells[0].value).upper() if column_cells[0].value else ""
                        is_rasatlar = "RASATLAR" in header_val
                        is_long_text = "AÇIKLAMA" in header_val or "MESAJ" in header_val

                        for cell in column_cells:
                            try:
                                if cell.value:
                                    cell_len = len(str(cell.value))
                                    if cell_len > max_length: max_length = cell_len
                                    # Uzun metinleri kaydır
                                    if is_rasatlar or is_long_text or cell_len > 50:
                                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                            except: pass
                        
                        if is_rasatlar:
                            ws.column_dimensions[col_letter].width = 35
                        elif is_long_text:
                            ws.column_dimensions[col_letter].width = 50
                        else:
                            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 60)

            messagebox.showinfo("Başarılı", f"İşlem Tamamlandı!\nDosya: {cikti}")
            
            # Okuma özetini göster
            if okuma_raporu:
                messagebox.showinfo("Okuma Özeti", okuma_raporu)

            try: os.startfile(cikti)
            except: pass

        except Exception as e:
            messagebox.showerror("Hata", f"Bir hata oluştu:\n{e}")
        finally:
            btn_run.config(state=tk.NORMAL, text=get_default_btn_text())

    threading.Thread(target=islem_yurut).start()

def get_default_btn_text():
    return "RAPOR OLUŞTUR"

root = tk.Tk()
root.title("Meteoroloji Denetim")
root.geometry("300x150")

btn_run = tk.Button(root, text=get_default_btn_text(), command=aylik_rapor_olustur, font=("Arial", 12, "bold"), bg="green", fg="white")
btn_run.pack(expand=True, fill="both", padx=20, pady=20)

root.mainloop()